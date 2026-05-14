from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy.orm import Session
from pydantic import BaseModel
from typing import Optional
from datetime import date
from fastapi.responses import StreamingResponse
import openpyxl
import io
from app.database import get_db
from app.models import User
from app.auth import get_current_user
from app.hr_models import HRHoliday

router = APIRouter()

class HolidayCreate(BaseModel):
    name: str
    date: date
    branch_id: Optional[int] = None
    holiday_type: str = "national"  # national / company

class HolidayUpdate(BaseModel):
    name: Optional[str] = None
    date: Optional[date] = None
    branch_id: Optional[int] = None
    holiday_type: Optional[str] = None
    is_active: Optional[bool] = None

def serialize(h: HRHoliday):
    return {
        "id": h.id,
        "name": h.name,
        "date": str(h.date),
        "branch_id": h.branch_id,
        "branch_name": h.branch.name if h.branch else "All Branches",
        "holiday_type": h.holiday_type,
        "is_active": h.is_active,
    }

@router.get("/")
def list_holidays(
    branch_id: Optional[int] = None,
    year: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    from sqlalchemy import extract
    q = db.query(HRHoliday).filter(HRHoliday.is_active == True)
    if branch_id:
        q = q.filter((HRHoliday.branch_id == branch_id) | (HRHoliday.branch_id == None))
    if year:
        q = q.filter(extract('year', HRHoliday.date) == year)
    return [serialize(h) for h in q.order_by(HRHoliday.date).all()]

@router.post("/")
def create_holiday(data: HolidayCreate, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    holiday = HRHoliday(**data.model_dump())
    db.add(holiday); db.commit(); db.refresh(holiday)
    return serialize(holiday)

@router.put("/{holiday_id}")
def update_holiday(holiday_id: int, data: HolidayUpdate, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    holiday = db.query(HRHoliday).filter(HRHoliday.id == holiday_id).first()
    if not holiday: raise HTTPException(404, "Holiday not found")
    for k, v in data.model_dump(exclude_none=True).items():
        setattr(holiday, k, v)
    db.commit(); db.refresh(holiday)
    return serialize(holiday)

@router.delete("/{holiday_id}")
def delete_holiday(holiday_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    holiday = db.query(HRHoliday).filter(HRHoliday.id == holiday_id).first()
    if not holiday: raise HTTPException(404, "Holiday not found")
    db.delete(holiday); db.commit()
    return {"message": "Deleted"}

@router.get("/import/template")
def get_holiday_template(current_user: User = Depends(get_current_user)):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Holiday Import Template"
    headers = ["name", "date", "branch_id", "holiday_type"]
    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=h)
    # Sample row
    ws.cell(row=2, column=1, value="New Year")
    ws.cell(row=2, column=2, value="2024-01-01")
    ws.cell(row=2, column=3, value="")
    ws.cell(row=2, column=4, value="national")
    
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=holiday_import_template.xlsx"}
    )

@router.post("/import/excel")
async def import_holidays_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    try:
        contents = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
        ws = wb.active
        
        imported = 0
        errors = []
        
        def safe_date(val):
            if isinstance(val, (date, datetime)): return val
            if val is None or str(val).strip() == "": return None
            try:
                for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y"):
                    try: return datetime.strptime(str(val).strip(), fmt).date()
                    except: continue
                return None
            except: return None

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            if not row[0] or str(row[0]).strip() == "":
                continue
                
            try:
                holiday_date = safe_date(row[1])
                if not holiday_date:
                    errors.append(f"Row {row_idx}: Invalid date format")
                    continue

                holiday = HRHoliday(
                    name=str(row[0]).strip(),
                    date=holiday_date,
                    branch_id=int(float(row[2])) if row[2] and str(row[2]).strip() != "" else None,
                    holiday_type=str(row[3]).strip() if row[3] else "national",
                    is_active=True
                )
                db.add(holiday)
                db.flush()
                imported += 1
            except Exception as e:
                db.rollback()
                errors.append(f"Row {row_idx}: {str(e)}")
                
        db.commit()
        return {"imported": imported, "errors": errors}
    except Exception as e:
        raise HTTPException(400, f"Critical failure: {str(e)}")


