from fastapi import APIRouter, Depends
from sqlalchemy.orm import Session
from sqlalchemy import func, extract
from typing import Optional
from datetime import date
from fastapi.responses import StreamingResponse
from app.database import get_db
from app.models import User
from app.auth import get_current_user
from app.hr_models import HRAttendanceRecord, HREmployee, HRLeaveBalance, HRLeaveType, HRPayrollRecord, HRArrearRecord
import io, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

router = APIRouter()

STATUS_LABEL = {
    "present": "P", "late": "L", "absent": "A", "half_day": "H",
    "leave": "LV", "on_duty": "OD", "holiday": "HOL", "weekly_off": "WO"
}

def _get_employees(db, branch_id=None, department_id=None, employee_id=None):
    q = db.query(HREmployee).filter(HREmployee.is_active == True)
    if branch_id: q = q.filter(HREmployee.branch_id == branch_id)
    if department_id: q = q.filter(HREmployee.department_id == department_id)
    if employee_id: q = q.filter(HREmployee.id == employee_id)
    return q.order_by(HREmployee.name).all()

@router.get("/monthly")
def monthly_attendance(
    month: int, year: int,
    branch_id: Optional[int] = None,
    department_id: Optional[int] = None,
    employee_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    employees = _get_employees(db, branch_id, department_id, employee_id)
    emp_ids = [e.id for e in employees]

    records = db.query(HRAttendanceRecord).filter(
        HRAttendanceRecord.employee_id.in_(emp_ids),
        extract('month', HRAttendanceRecord.date) == month,
        extract('year', HRAttendanceRecord.date) == year,
    ).all()

    record_map = {}
    for r in records:
        record_map[(r.employee_id, str(r.date))] = r.status

    from calendar import monthrange
    days_in_month = monthrange(year, month)[1]

    result = []
    for emp in employees:
        row = {
            "employee_id": emp.id,
            "employee_code": emp.employee_id,
            "employee_name": emp.name,
            "department": emp.department.name if emp.department else None,
            "days": {}
        }
        summary = {"present": 0, "late": 0, "absent": 0, "half_day": 0,
                   "leave": 0, "on_duty": 0, "holiday": 0, "weekly_off": 0}
        for d in range(1, days_in_month + 1):
            day_str = f"{year}-{str(month).zfill(2)}-{str(d).zfill(2)}"
            status = record_map.get((emp.id, day_str), "absent")
            row["days"][day_str] = {"status": status, "label": STATUS_LABEL.get(status, "?")}
            if status in summary:
                summary[status] += 1
        row["summary"] = summary
        result.append(row)

    return result

@router.get("/late-arrivals")
def late_arrivals(
    month: int, year: int,
    branch_id: Optional[int] = None,
    department_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    q = db.query(HRAttendanceRecord).filter(
        HRAttendanceRecord.is_late == True,
        extract('month', HRAttendanceRecord.date) == month,
        extract('year', HRAttendanceRecord.date) == year,
    )
    records = q.all()
    return [{
        "employee_id": r.employee_id,
        "employee_name": r.employee.name if r.employee else None,
        "date": str(r.date),
        "check_in": str(r.check_in) if r.check_in else None,
        "late_minutes": r.late_minutes,
    } for r in records]

@router.get("/absenteeism")
def absenteeism_report(
    month: int, year: int,
    branch_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    q = db.query(HRAttendanceRecord).filter(
        HRAttendanceRecord.status == "absent",
        extract('month', HRAttendanceRecord.date) == month,
        extract('year', HRAttendanceRecord.date) == year,
    )
    records = q.all()
    return [{
        "employee_id": r.employee_id,
        "employee_name": r.employee.name if r.employee else None,
        "date": str(r.date),
    } for r in records]

@router.get("/leave-summary")
def leave_summary(
    year: int,
    branch_id: Optional[int] = None,
    department_id: Optional[int] = None,
    employee_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    employees = _get_employees(db, branch_id, department_id, employee_id)
    leave_types = db.query(HRLeaveType).filter(HRLeaveType.is_active == True).all()
    result = []
    for emp in employees:
        row = {"employee_id": emp.id, "employee_name": emp.name, "leave_types": {}}
        for lt in leave_types:
            balance = db.query(HRLeaveBalance).filter(
                HRLeaveBalance.employee_id == emp.id,
                HRLeaveBalance.leave_type_id == lt.id,
                HRLeaveBalance.year == year
            ).first()
            row["leave_types"][lt.code] = {
                "allocated": balance.allocated_days if balance else 0,
                "used": balance.used_days if balance else 0,
                "remaining": (balance.allocated_days - balance.used_days) if balance else 0,
            }
        result.append(row)
    return result

@router.get("/payroll-export")
def payroll_export(
    month: int, year: int,
    branch_id: Optional[int] = None,
    department_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    q = db.query(HRPayrollRecord).join(HREmployee, HRPayrollRecord.employee_id == HREmployee.id).filter(
        HRPayrollRecord.month == month,
        HRPayrollRecord.year == year
    )
    if branch_id:
        q = q.filter(HREmployee.branch_id == branch_id)
    if department_id:
        q = q.filter(HREmployee.department_id == department_id)
        
    records = q.all()
    result = []
    for payroll in records:
        emp = payroll.employee
        if not emp: continue
        result.append({
            "employee_id": emp.employee_id,
            "employee_name": emp.name,
            "designation": emp.designation,
            "working_days": payroll.working_days,
            "present_days": payroll.present_days,
            "absent_days": payroll.absent_days,
            "leave_days": payroll.leave_days,
            "lop_days": payroll.lop_days,
            "on_duty_days": payroll.on_duty_days,
            "basic_salary": float(payroll.basic_salary or 0),
            "total_earnings": float(payroll.total_earnings or 0),
            "total_deductions": float(payroll.total_deductions or 0),
            "net_salary": float(payroll.net_salary or 0),
            "status": payroll.status,
        })
    return result

@router.get("/payroll-export/excel")
def payroll_export_excel(
    month: int, year: int,
    branch_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    q = db.query(HRPayrollRecord).join(HREmployee, HRPayrollRecord.employee_id == HREmployee.id).filter(
        HRPayrollRecord.month == month,
        HRPayrollRecord.year == year
    )
    if branch_id:
        q = q.filter(HREmployee.branch_id == branch_id)
    payroll_records = q.all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Payroll {month}-{year}"

    if not payroll_records:
        headers = ["Emp ID", "Name", "Designation", "Working Days", "Present", "Absent", "Leave", "LOP", "On Duty", "Basic", "Earnings", "Deductions", "Net Salary"]
        header_fill = PatternFill("solid", fgColor="1a472a")
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
    else:
        emp_map = {pr.employee_id: pr.employee for pr in payroll_records if pr.employee}
        all_regular_earning_keys = set()
        all_arrear_earning_keys = set()
        all_deduction_keys = set()
        all_employer_cont_keys = set()

        for pr in payroll_records:
            breakdown = pr.components_breakdown or {}
            earnings = breakdown.get('earnings', {})
            deductions = breakdown.get('deductions', {})
            employer_cont = breakdown.get('employer_contributions', {})

            for k in earnings.keys():
                if 'Arrear' in k:
                    all_arrear_earning_keys.add(k)
                else:
                    all_regular_earning_keys.add(k)
            for k in deductions.keys():
                all_deduction_keys.add(k)
            for k in employer_cont.keys():
                all_employer_cont_keys.add(k)

        regular_earnings_cols = sorted(list(all_regular_earning_keys))
        # Keep "Basic Salary" or "Basic" first in regular earnings if present
        for b_key in ["Basic Salary", "Basic"]:
            if b_key in regular_earnings_cols:
                regular_earnings_cols.remove(b_key)
                regular_earnings_cols.insert(0, b_key)

        arrear_earnings_cols = sorted(list(all_arrear_earning_keys))
        deductions_cols = sorted(list(all_deduction_keys))
        employer_cont_cols = sorted(list(all_employer_cont_keys))

        # Build Header List
        headers = [
            "Emp ID", "Name", "Designation", "Department", "UAN", "ESI Number",
            "Working Days", "Present", "Absent", "Leave", "LOP", "On Duty", "Week Off"
        ]
        headers.extend(regular_earnings_cols)
        headers.append("Gross Earnings")
        headers.append("Arrears Paid")
        headers.append("Arrears Paid Details")

        if arrear_earnings_cols:
            headers.extend(arrear_earnings_cols)
            headers.append("Gross with Arrears")

        headers.extend(deductions_cols)
        headers.append("Total Deductions")
        headers.append("Arrears Held")
        headers.append("Arrears Held Details")

        headers.extend(employer_cont_cols)
        headers.append("Total Employer Contributions")

        headers.append("Monthly CTC")
        headers.append("Net Salary")
        headers.append("Total Pending Arrears")
        headers.append("Pending Arrears Balance")

        header_fill = PatternFill("solid", fgColor="1a472a")
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        row_num = 2
        for pr in payroll_records:
            emp = emp_map.get(pr.employee_id)
            if not emp:
                continue
            breakdown = pr.components_breakdown or {}
            earnings = breakdown.get('earnings', {})
            deductions = breakdown.get('deductions', {})
            employer_cont = breakdown.get('employer_contributions', {})

            # Count week off days
            import calendar
            days_in_month = calendar.monthrange(year, month)[1]
            
            has_attendance = db.query(HRAttendanceRecord).filter(
                HRAttendanceRecord.employee_id == pr.employee_id,
                HRAttendanceRecord.date >= date(year, month, 1),
                HRAttendanceRecord.date <= date(year, month, days_in_month)
            ).first() is not None
            
            if has_attendance:
                week_offs = db.query(HRAttendanceRecord).filter(
                    HRAttendanceRecord.employee_id == pr.employee_id,
                    HRAttendanceRecord.date >= date(year, month, 1),
                    HRAttendanceRecord.date <= date(year, month, days_in_month),
                    HRAttendanceRecord.status == "weekly_off"
                ).count()
            else:
                from app.routers.hr_payroll import get_hr_config
                global_working_days = get_hr_config(db, "working_days", ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat"])
                DAY_MAP = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
                week_offs = 0
                for d_day in range(1, days_in_month + 1):
                    dt_day = date(year, month, d_day)
                    if DAY_MAP[dt_day.weekday()] not in global_working_days:
                        week_offs += 1

            row_data = {
                "Emp ID": emp.employee_id,
                "Name": emp.name,
                "Designation": emp.designation,
                "Department": emp.department.name if emp.department else "-",
                "UAN": emp.uan or "-",
                "ESI Number": emp.esi_number or "-",
                "Working Days": pr.working_days,
                "Present": pr.present_days,
                "Absent": pr.absent_days,
                "Leave": pr.leave_days,
                "LOP": pr.lop_days,
                "On Duty": pr.on_duty_days,
                "Week Off": week_offs,
            }

            # Populate regular earnings
            for col_name in regular_earnings_cols:
                row_data[col_name] = float(earnings.get(col_name, 0))
            regular_gross = sum(float(v) for k, v in earnings.items() if 'Arrear' not in k)
            row_data["Gross Earnings"] = regular_gross
            row_data["Arrears Paid"] = float(pr.arrears_paid or 0)

            # Query all arrear records for this employee to build details
            all_arrs = db.query(HRArrearRecord).filter(HRArrearRecord.employee_id == pr.employee_id).all()

            # 1. Arrears Paid Details
            paid_items = []
            for a in all_arrs:
                if a.status == "paid" and a.paid_in_month == month and a.paid_in_year == year:
                    rem = a.remarks or "Paid"
                    paid_items.append(f"{rem} ({a.held_month}/{a.held_year}): ₹{a.amount_held}")
            row_data["Arrears Paid Details"] = "; ".join(paid_items) if paid_items else "-"

            # Populate arrears
            if arrear_earnings_cols:
                for col_name in arrear_earnings_cols:
                    row_data[col_name] = float(earnings.get(col_name, 0))
                row_data["Gross with Arrears"] = float(pr.total_earnings or 0)

            # Populate deductions
            for col_name in deductions_cols:
                row_data[col_name] = float(deductions.get(col_name, 0))
            row_data["Total Deductions"] = float(pr.total_deductions or 0)
            row_data["Arrears Held"] = float(pr.arrears_held or 0)

            # 2. Arrears Held Details
            held_items = []
            for a in all_arrs:
                # Newly held
                is_new_hold = (a.status == "held" and a.held_month == month and a.held_year == year)
                # Deducted (collected from salary)
                is_deducted = (a.status == "deducted" and a.paid_in_month == month and a.paid_in_year == year) or \
                              (a.status == "one_time" and a.held_month == month and a.held_year == year)
                if is_new_hold or is_deducted:
                    rem = a.remarks or "Held/Deducted"
                    held_items.append(f"{rem} ({a.held_month}/{a.held_year}): ₹{a.amount_held}")
            row_data["Arrears Held Details"] = "; ".join(held_items) if held_items else "-"

            # Populate employer contributions
            for col_name in employer_cont_cols:
                row_data[col_name] = float(employer_cont.get(col_name, 0))
            total_employer_cont = sum(float(v) for v in employer_cont.values())
            row_data["Total Employer Contributions"] = total_employer_cont

            # Populate Monthly CTC and Net Salary
            row_data["Monthly CTC"] = regular_gross + total_employer_cont
            row_data["Net Salary"] = float(round(pr.net_salary or 0))

            # 3. Pending Arrears Balance & Total Pending Arrears
            pending_items = []
            total_pending_amt = 0.0
            for a in all_arrs:
                if a.status in ["held", "one_time"] and float(a.amount_held or 0) > 0:
                    amt = float(a.amount_held or 0)
                    total_pending_amt += amt
                    rem = a.remarks or "Pending"
                    pending_items.append(f"{rem} ({a.held_month}/{a.held_year}): ₹{a.amount_held}")
            row_data["Total Pending Arrears"] = total_pending_amt
            row_data["Pending Arrears Balance"] = "; ".join(pending_items) if pending_items else "-"

            # Write values to cells
            for col_num, header in enumerate(headers, 1):
                val = row_data.get(header, "")
                cell = ws.cell(row=row_num, column=col_num, value=val)
                if isinstance(val, (int, float)) and header not in ["Emp ID", "Working Days", "Present", "Absent", "Leave", "LOP", "On Duty", "Week Off"]:
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal="right")
            row_num += 1

    # --- Add Arrears Tracker Sheet ---
    ws_arrears = wb.create_sheet(title="Arrears Tracker")
    
    arrear_headers = ["Emp ID", "Name", "Designation", "Department", "Held Period (M/Y)", "Amount", "Status", "Paid/Deducted Period (M/Y)", "Remarks"]
    arrear_header_fill = PatternFill("solid", fgColor="d97706") # Amber header to match Arrears styling
    
    for col_num, h in enumerate(arrear_headers, 1):
        cell = ws_arrears.cell(row=1, column=col_num, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = arrear_header_fill
        cell.alignment = Alignment(horizontal="center")
        
    arrear_records = db.query(HRArrearRecord).join(HREmployee, HRArrearRecord.employee_id == HREmployee.id).order_by(
        HREmployee.name, 
        HRArrearRecord.held_year.desc(), 
        HRArrearRecord.held_month.desc()
    ).all()
    
    a_row_num = 2
    for arr in arrear_records:
        emp = arr.employee
        if not emp:
            continue
            
        held_period = f"{arr.held_month}/{arr.held_year}" if arr.held_month else "-"
        paid_period = f"{arr.paid_in_month}/{arr.paid_in_year}" if arr.paid_in_month else "-"
        status_text = arr.status.upper()
        
        ws_arrears.cell(row=a_row_num, column=1, value=emp.employee_id)
        ws_arrears.cell(row=a_row_num, column=2, value=emp.name)
        ws_arrears.cell(row=a_row_num, column=3, value=emp.designation or "-")
        ws_arrears.cell(row=a_row_num, column=4, value=emp.department.name if emp.department else "-")
        ws_arrears.cell(row=a_row_num, column=5, value=held_period)
        
        cell_amt = ws_arrears.cell(row=a_row_num, column=6, value=float(arr.amount_held or 0))
        cell_amt.number_format = '#,##0.00'
        cell_amt.alignment = Alignment(horizontal="right")
        
        ws_arrears.cell(row=a_row_num, column=7, value=status_text)
        ws_arrears.cell(row=a_row_num, column=8, value=paid_period)
        ws_arrears.cell(row=a_row_num, column=9, value=arr.remarks or "")
        
        a_row_num += 1

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    filename = f"payroll_{year}_{str(month).zfill(2)}.xlsx"
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f"attachment; filename={filename}"})
