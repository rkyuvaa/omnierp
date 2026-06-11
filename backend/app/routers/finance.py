"""
Finance Router — /api/finance
Full-featured finance module: account heads, budgets, bank statement import,
weekly buckets, pivot report, and management summary.
"""
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func, and_, or_
from typing import Optional, List
from datetime import datetime, date, timedelta
from pydantic import BaseModel
import io, csv, calendar

from app.database import get_db
from app.models import User
from app.auth import get_current_user
from app.bank_models import (
    BankAccount, BankTransaction, AccountHead, BudgetMaster,
    MappingRule, BankColumnMapping, ImportBatch
)

router = APIRouter()

# ── Auth helpers ──────────────────────────────────────────────────────────────
def require_login(current_user: User = Depends(get_current_user)) -> User:
    if not current_user:
        raise HTTPException(401, "Not authenticated")
    return current_user

def require_admin(current_user: User = Depends(get_current_user)) -> User:
    if not current_user or not current_user.is_superadmin:
        raise HTTPException(403, "Admin access required")
    return current_user

# ── Utility: Week calculation (Sunday-start) ─────────────────────────────────
def get_week_info(d: date) -> tuple:
    """Return (week_number, week_year, week_start_sunday, week_end_saturday)."""
    # Python: weekday() Mon=0 … Sun=6
    days_since_sunday = (d.weekday() + 1) % 7
    sunday = d - timedelta(days=days_since_sunday)
    saturday = sunday + timedelta(days=6)
    # %U = week number with Sunday as first day (00–53)
    week_num = int(d.strftime('%U'))
    return week_num, d.year, sunday, saturday


def weeks_in_month(year: int, month: int) -> List[dict]:
    """Return all Sun-Sat weeks that overlap with the given month."""
    first_day = date(year, month, 1)
    last_day = date(year, month, calendar.monthrange(year, month)[1])
    weeks = []
    seen = set()
    d = first_day
    while d <= last_day:
        wn, wy, ws, we = get_week_info(d)
        key = (wy, wn)
        if key not in seen:
            seen.add(key)
            weeks.append({"week_num": wn, "week_year": wy, "start": ws, "end": we})
        d += timedelta(days=1)
    return weeks


def current_fy() -> str:
    """Return Indian financial year string like '2025-26'."""
    today = date.today()
    if today.month >= 4:
        return f"{today.year}-{str(today.year + 1)[2:]}"
    return f"{today.year - 1}-{str(today.year)[2:]}"


# ── Reference generator ───────────────────────────────────────────────────────
def _next_ref(db: Session) -> str:
    count = db.query(func.count(BankTransaction.id)).scalar() or 0
    return f"PAY-{count + 1:04d}"

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# SERIALIZERS
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def _ser_account(a: BankAccount) -> dict:
    bal = float(a.last_statement_balance or a.current_balance or 0)
    lien = float(a.lien_amount or 0)
    return {
        "id": a.id,
        "account_name": a.account_name,
        "account_number": a.account_number,
        "bank_name": a.bank_name,
        "branch_name": a.branch_name,
        "ifsc_code": a.ifsc_code,
        "opening_balance": float(a.opening_balance or 0),
        "current_balance": float(a.current_balance or 0),
        "last_statement_balance": float(a.last_statement_balance or 0) if a.last_statement_balance is not None else None,
        "lien_amount": lien,
        "actual_balance": round(bal - lien, 2),
        "last_import_at": a.last_import_at.isoformat() if a.last_import_at else None,
        "is_active": a.is_active,
        "notes": a.notes,
        "created_at": a.created_at.isoformat() if a.created_at else None,
        "updated_at": a.updated_at.isoformat() if a.updated_at else None,
    }


def _ser_txn(t: BankTransaction) -> dict:
    return {
        "id": t.id,
        "reference": t.reference,
        "account_id": t.account_id,
        "transaction_date": t.transaction_date.isoformat() if t.transaction_date else None,
        "payee_name": t.payee_name,
        "description": t.description,
        "amount": float(t.amount or 0),
        "debit_amount": float(t.debit_amount or 0),
        "credit_amount": float(t.credit_amount or 0),
        "transaction_type": t.transaction_type or "DEBIT",
        "account_head_id": t.account_head_id,
        "account_head_name": t.account_head.head_name if t.account_head else None,
        "account_head_code": t.account_head.head_code if t.account_head else None,
        "import_batch_id": t.import_batch_id,
        "week_number": t.week_number,
        "week_year": t.week_year,
        "status": t.status,
        "entered_by": t.entered_by,
        "entered_by_name": t.entered_by_name,
        "entered_at": t.entered_at.isoformat() if t.entered_at else None,
        "verified_by": t.verified_by,
        "verified_by_name": t.verified_by_name,
        "verified_at": t.verified_at.isoformat() if t.verified_at else None,
        "verification_remarks": t.verification_remarks,
        "approved_by": t.approved_by,
        "approved_by_name": t.approved_by_name,
        "approved_at": t.approved_at.isoformat() if t.approved_at else None,
        "approval_remarks": t.approval_remarks,
        "rejection_remarks": t.rejection_remarks,
        "created_at": t.created_at.isoformat() if t.created_at else None,
        "updated_at": t.updated_at.isoformat() if t.updated_at else None,
    }


def _ser_head(h: AccountHead) -> dict:
    return {
        "id": h.id,
        "head_code": h.head_code,
        "head_name": h.head_name,
        "category": h.category,
        "is_active": h.is_active,
        "created_at": h.created_at.isoformat() if h.created_at else None,
    }


def _ser_budget(b: BudgetMaster) -> dict:
    return {
        "id": b.id,
        "financial_year": b.financial_year,
        "month": b.month,
        "account_head_id": b.account_head_id,
        "head_code": b.account_head.head_code if b.account_head else None,
        "head_name": b.account_head.head_name if b.account_head else None,
        "planned_amount": float(b.planned_amount or 0),
        "updated_at": b.updated_at.isoformat() if b.updated_at else None,
    }


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# PYDANTIC SCHEMAS
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

class AccountCreate(BaseModel):
    account_name: str
    account_number: Optional[str] = None
    bank_name: Optional[str] = None
    branch_name: Optional[str] = None
    ifsc_code: Optional[str] = None
    opening_balance: Optional[float] = 0
    lien_amount: Optional[float] = 0
    notes: Optional[str] = None

class AccountUpdate(BaseModel):
    account_name: Optional[str] = None
    account_number: Optional[str] = None
    bank_name: Optional[str] = None
    branch_name: Optional[str] = None
    ifsc_code: Optional[str] = None
    opening_balance: Optional[float] = None
    lien_amount: Optional[float] = None
    notes: Optional[str] = None
    is_active: Optional[bool] = None

class HeadCreate(BaseModel):
    head_code: str
    head_name: str
    category: Optional[str] = None

class HeadUpdate(BaseModel):
    head_name: Optional[str] = None
    category: Optional[str] = None
    is_active: Optional[bool] = None

class BudgetUpsert(BaseModel):
    financial_year: str
    month: int
    account_head_id: int
    planned_amount: float

class MappingRuleCreate(BaseModel):
    keyword: str
    account_head_id: int
    bank_account_id: Optional[int] = None

class ColumnMappingUpsert(BaseModel):
    bank_account_id: int
    date_col: str = "Date"
    description_col: str = "Description"
    debit_col: str = "Debit"
    credit_col: str = "Credit"
    balance_col: str = "Balance"
    date_format: str = "%d/%m/%Y"
    skip_rows: int = 0

class ImportRow(BaseModel):
    transaction_date: Optional[str] = None
    description: Optional[str] = None
    debit_amount: Optional[float] = 0
    credit_amount: Optional[float] = 0
    balance: Optional[float] = None
    account_head_id: Optional[int] = None
    transaction_type: str = "DEBIT"

class ImportPostRequest(BaseModel):
    bank_account_id: int
    filename: Optional[str] = None
    rows: List[ImportRow]

class BulkActionRequest(BaseModel):
    transaction_ids: List[int]
    remarks: Optional[str] = None

class TransactionActionRequest(BaseModel):
    remarks: Optional[str] = None

class TransactionCreate(BaseModel):
    account_id: int
    transaction_date: Optional[date] = None
    payee_name: Optional[str] = None
    description: Optional[str] = None
    amount: float
    account_head_id: Optional[int] = None


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# BANK ACCOUNTS
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

@router.get("/accounts")
def list_accounts(include_inactive: bool = False, db: Session = Depends(get_db), current_user: User = Depends(require_login)):
    q = db.query(BankAccount)
    if not include_inactive:
        q = q.filter(BankAccount.is_active == True)
    return [_ser_account(a) for a in q.all()]


@router.post("/accounts")
def create_account(data: AccountCreate, db: Session = Depends(get_db), current_user: User = Depends(require_admin)):
    acc = BankAccount(
        account_name=data.account_name,
        account_number=data.account_number,
        bank_name=data.bank_name,
        branch_name=data.branch_name,
        ifsc_code=data.ifsc_code,
        opening_balance=data.opening_balance or 0,
        current_balance=data.opening_balance or 0,
        lien_amount=data.lien_amount or 0,
        notes=data.notes,
        created_by=current_user.id,
    )
    db.add(acc)
    db.commit()
    db.refresh(acc)
    return _ser_account(acc)


@router.put("/accounts/{acc_id}")
def update_account(acc_id: int, data: AccountUpdate, db: Session = Depends(get_db), current_user: User = Depends(require_admin)):
    acc = db.query(BankAccount).filter(BankAccount.id == acc_id).first()
    if not acc:
        raise HTTPException(404, "Account not found")
    for k, v in data.model_dump(exclude_none=True).items():
        setattr(acc, k, v)
    acc.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(acc)
    return _ser_account(acc)


@router.delete("/accounts/{acc_id}")
def delete_account(acc_id: int, db: Session = Depends(get_db), current_user: User = Depends(require_admin)):
    acc = db.query(BankAccount).filter(BankAccount.id == acc_id).first()
    if not acc:
        raise HTTPException(404, "Account not found")
    acc.is_active = False
    acc.updated_at = datetime.utcnow()
    db.commit()
    return {"message": "Account deactivated"}


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# ACCOUNT HEADS
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

@router.get("/heads")
def list_heads(include_inactive: bool = False, db: Session = Depends(get_db), current_user: User = Depends(require_login)):
    q = db.query(AccountHead)
    if not include_inactive:
        q = q.filter(AccountHead.is_active == True)
    return [_ser_head(h) for h in q.order_by(AccountHead.head_code).all()]


@router.post("/heads")
def create_head(data: HeadCreate, db: Session = Depends(get_db), current_user: User = Depends(require_admin)):
    existing = db.query(AccountHead).filter(AccountHead.head_code == data.head_code.upper()).first()
    if existing:
        raise HTTPException(400, f"Head code {data.head_code} already exists")
    h = AccountHead(head_code=data.head_code.upper(), head_name=data.head_name, category=data.category)
    db.add(h)
    db.commit()
    db.refresh(h)
    return _ser_head(h)


@router.put("/heads/{head_id}")
def update_head(head_id: int, data: HeadUpdate, db: Session = Depends(get_db), current_user: User = Depends(require_admin)):
    h = db.query(AccountHead).filter(AccountHead.id == head_id).first()
    if not h:
        raise HTTPException(404, "Account head not found")
    for k, v in data.model_dump(exclude_none=True).items():
        setattr(h, k, v)
    h.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(h)
    return _ser_head(h)


@router.delete("/heads/{head_id}")
def delete_head(head_id: int, db: Session = Depends(get_db), current_user: User = Depends(require_admin)):
    h = db.query(AccountHead).filter(AccountHead.id == head_id).first()
    if not h:
        raise HTTPException(404, "Account head not found")
    h.is_active = False
    h.updated_at = datetime.utcnow()
    db.commit()
    return {"message": "Account head deactivated"}


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# BUDGET MASTER
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

@router.get("/budgets")
def list_budgets(
    fy: Optional[str] = None,
    month: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_login)
):
    q = db.query(BudgetMaster)
    if fy:
        q = q.filter(BudgetMaster.financial_year == fy)
    if month:
        q = q.filter(BudgetMaster.month == month)
    return [_ser_budget(b) for b in q.join(BudgetMaster.account_head).order_by(BudgetMaster.month, AccountHead.head_code).all()]


@router.post("/budgets")
def upsert_budget(data: BudgetUpsert, db: Session = Depends(get_db), current_user: User = Depends(require_admin)):
    b = db.query(BudgetMaster).filter(
        BudgetMaster.financial_year == data.financial_year,
        BudgetMaster.month == data.month,
        BudgetMaster.account_head_id == data.account_head_id,
    ).first()
    if b:
        b.planned_amount = data.planned_amount
        b.updated_at = datetime.utcnow()
    else:
        b = BudgetMaster(
            financial_year=data.financial_year,
            month=data.month,
            account_head_id=data.account_head_id,
            planned_amount=data.planned_amount,
        )
        db.add(b)
    db.commit()
    db.refresh(b)
    return _ser_budget(b)


@router.delete("/budgets/{budget_id}")
def delete_budget(budget_id: int, db: Session = Depends(get_db), current_user: User = Depends(require_admin)):
    b = db.query(BudgetMaster).filter(BudgetMaster.id == budget_id).first()
    if not b:
        raise HTTPException(404, "Budget not found")
    db.delete(b)
    db.commit()
    return {"message": "Budget deleted"}


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# MAPPING RULES
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

@router.get("/mapping-rules")
def list_mapping_rules(bank_id: Optional[int] = None, db: Session = Depends(get_db), current_user: User = Depends(require_login)):
    q = db.query(MappingRule).filter(MappingRule.is_active == True)
    if bank_id:
        q = q.filter(or_(MappingRule.bank_account_id == bank_id, MappingRule.bank_account_id == None))
    rules = q.all()
    return [{
        "id": r.id,
        "keyword": r.keyword,
        "account_head_id": r.account_head_id,
        "head_name": r.account_head.head_name if r.account_head else None,
        "head_code": r.account_head.head_code if r.account_head else None,
        "bank_account_id": r.bank_account_id,
        "bank_name": r.bank_account.account_name if r.bank_account else None,
        "created_at": r.created_at.isoformat() if r.created_at else None,
    } for r in rules]


@router.post("/mapping-rules")
def create_mapping_rule(data: MappingRuleCreate, db: Session = Depends(get_db), current_user: User = Depends(require_login)):
    r = MappingRule(
        keyword=data.keyword,
        account_head_id=data.account_head_id,
        bank_account_id=data.bank_account_id,
    )
    db.add(r)
    db.commit()
    db.refresh(r)
    return {"id": r.id, "keyword": r.keyword, "account_head_id": r.account_head_id}


@router.delete("/mapping-rules/{rule_id}")
def delete_mapping_rule(rule_id: int, db: Session = Depends(get_db), current_user: User = Depends(require_login)):
    r = db.query(MappingRule).filter(MappingRule.id == rule_id).first()
    if not r:
        raise HTTPException(404, "Rule not found")
    db.delete(r)
    db.commit()
    return {"message": "Rule deleted"}


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# BANK COLUMN MAPPING (per-bank CSV/Excel config)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

@router.get("/column-mapping/{bank_account_id}")
def get_column_mapping(bank_account_id: int, db: Session = Depends(get_db), current_user: User = Depends(require_login)):
    cm = db.query(BankColumnMapping).filter(BankColumnMapping.bank_account_id == bank_account_id).first()
    if not cm:
        return {"bank_account_id": bank_account_id, "date_col": "Date", "description_col": "Description",
                "debit_col": "Debit", "credit_col": "Credit", "balance_col": "Balance",
                "date_format": "%d/%m/%Y", "skip_rows": 0}
    return {
        "id": cm.id,
        "bank_account_id": cm.bank_account_id,
        "date_col": cm.date_col,
        "description_col": cm.description_col,
        "debit_col": cm.debit_col,
        "credit_col": cm.credit_col,
        "balance_col": cm.balance_col,
        "date_format": cm.date_format,
        "skip_rows": cm.skip_rows,
    }


@router.post("/column-mapping")
def upsert_column_mapping(data: ColumnMappingUpsert, db: Session = Depends(get_db), current_user: User = Depends(require_admin)):
    cm = db.query(BankColumnMapping).filter(BankColumnMapping.bank_account_id == data.bank_account_id).first()
    if cm:
        cm.date_col = data.date_col
        cm.description_col = data.description_col
        cm.debit_col = data.debit_col
        cm.credit_col = data.credit_col
        cm.balance_col = data.balance_col
        cm.date_format = data.date_format
        cm.skip_rows = data.skip_rows
        cm.updated_at = datetime.utcnow()
    else:
        cm = BankColumnMapping(
            bank_account_id=data.bank_account_id,
            date_col=data.date_col,
            description_col=data.description_col,
            debit_col=data.debit_col,
            credit_col=data.credit_col,
            balance_col=data.balance_col,
            date_format=data.date_format,
            skip_rows=data.skip_rows,
        )
        db.add(cm)
    db.commit()
    return {"message": "Column mapping saved"}


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# BANK STATEMENT IMPORT
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def _parse_amount(val) -> float:
    if val is None:
        return 0.0
    s = str(val).strip().replace(",", "").replace("₹", "").replace(" ", "")
    if s in ("", "-", "–"):
        return 0.0
    try:
        return abs(float(s))
    except:
        return 0.0


def _parse_date(val, fmt: str) -> Optional[str]:
    if not val:
        return None
    s = str(val).strip()
    for f in [fmt, "%d-%m-%Y", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d %b %Y", "%d-%b-%Y"]:
        try:
            return datetime.strptime(s, f).date().isoformat()
        except:
            continue
    return None


def _auto_map(description: str, rules: list, bank_account_id: int) -> Optional[int]:
    """Apply mapping rules. Bank-specific rules take precedence."""
    desc = (description or "").lower()
    bank_specific = [r for r in rules if r.bank_account_id == bank_account_id]
    global_rules = [r for r in rules if r.bank_account_id is None]
    for r in bank_specific + global_rules:
        if r.keyword.lower() in desc:
            return r.account_head_id
    return None


@router.post("/import/preview")
async def import_preview(
    bank_account_id: int = Form(...),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_login),
):
    """Parse uploaded CSV/Excel and return preview rows with auto-mapped account heads."""
    # Get column mapping for this bank
    cm = db.query(BankColumnMapping).filter(BankColumnMapping.bank_account_id == bank_account_id).first()
    date_col = cm.date_col if cm else "Date"
    desc_col = cm.description_col if cm else "Description"
    debit_col = cm.debit_col if cm else "Debit"
    credit_col = cm.credit_col if cm else "Credit"
    balance_col = cm.balance_col if cm else "Balance"
    date_fmt = cm.date_format if cm else "%d/%m/%Y"
    skip_rows = cm.skip_rows if cm else 0

    content = await file.read()
    filename = file.filename or ""
    rows = []

    try:
        if filename.lower().endswith((".xlsx", ".xls")):
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
            ws = wb.active
            all_rows = list(ws.iter_rows(values_only=True))
            # Skip rows + find header
            header_row = all_rows[skip_rows] if len(all_rows) > skip_rows else []
            headers = [str(h).strip() if h is not None else "" for h in header_row]
            data_rows = all_rows[skip_rows + 1:]
            for row in data_rows:
                if all(v is None for v in row):
                    continue
                row_dict = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
                rows.append(row_dict)
        else:
            # CSV
            text = content.decode("utf-8-sig", errors="replace")
            reader = csv.DictReader(io.StringIO(text))
            # skip_rows
            for i, row in enumerate(reader):
                if i < skip_rows:
                    continue
                rows.append(row)
    except Exception as e:
        raise HTTPException(400, f"Failed to parse file: {e}")

    # Get mapping rules
    rules = db.query(MappingRule).filter(MappingRule.is_active == True).all()

    preview = []
    for r in rows:
        raw_date = r.get(date_col) or r.get("Date") or r.get("Txn Date") or r.get("Transaction Date")
        raw_desc = r.get(desc_col) or r.get("Description") or r.get("Narration") or r.get("Particulars") or ""
        raw_debit = r.get(debit_col) or r.get("Debit") or r.get("DR") or ""
        raw_credit = r.get(credit_col) or r.get("Credit") or r.get("CR") or ""
        raw_balance = r.get(balance_col) or r.get("Balance") or ""

        debit_amt = _parse_amount(raw_debit)
        credit_amt = _parse_amount(raw_credit)

        if debit_amt == 0 and credit_amt == 0:
            continue  # skip empty rows

        txn_type = "CREDIT" if credit_amt > 0 and debit_amt == 0 else "DEBIT"
        auto_head = _auto_map(str(raw_desc), rules, bank_account_id) if txn_type == "DEBIT" else None

        preview.append({
            "transaction_date": _parse_date(raw_date, date_fmt),
            "description": str(raw_desc).strip(),
            "debit_amount": debit_amt,
            "credit_amount": credit_amt,
            "balance": _parse_amount(raw_balance),
            "transaction_type": txn_type,
            "account_head_id": auto_head,
            "auto_mapped": auto_head is not None,
        })

    # Return available column names from the file for mapping UI feedback
    sample_cols = list(rows[0].keys()) if rows else []

    return {"rows": preview, "available_columns": sample_cols, "filename": filename}


@router.post("/import/post")
def import_post(data: ImportPostRequest, db: Session = Depends(get_db), current_user: User = Depends(require_login)):
    """Confirm and post all import rows into transactions table."""
    acc = db.query(BankAccount).filter(BankAccount.id == data.bank_account_id, BankAccount.is_active == True).first()
    if not acc:
        raise HTTPException(404, "Bank account not found")

    # Create import batch
    batch = ImportBatch(
        bank_account_id=data.bank_account_id,
        uploaded_by=current_user.id,
        filename=data.filename,
        row_count=0,
        debit_count=0,
        credit_count=0,
        duplicate_skipped=0,
    )
    db.add(batch)
    db.flush()

    inserted = 0
    skipped = 0
    last_balance = None

    for row in data.rows:
        # Duplicate check: same bank + date + description + debit/credit amount
        txn_date = None
        if row.transaction_date:
            try:
                txn_date = date.fromisoformat(row.transaction_date)
            except:
                pass

        if txn_date:
            check_amount = row.debit_amount if row.transaction_type == "DEBIT" else row.credit_amount
            dup = db.query(BankTransaction).filter(
                BankTransaction.account_id == data.bank_account_id,
                BankTransaction.transaction_date == txn_date,
                BankTransaction.description == row.description,
            ).first()
            if dup:
                if row.transaction_type == "DEBIT" and abs(float(dup.debit_amount or dup.amount or 0) - check_amount) < 0.01:
                    skipped += 1
                    continue
                if row.transaction_type == "CREDIT" and abs(float(dup.credit_amount or 0) - check_amount) < 0.01:
                    skipped += 1
                    continue

        wn, wy, _, _ = get_week_info(txn_date) if txn_date else (None, None, None, None)
        main_amount = row.debit_amount if row.transaction_type == "DEBIT" else row.credit_amount

        txn = BankTransaction(
            reference=_next_ref(db),
            account_id=data.bank_account_id,
            transaction_date=txn_date,
            description=row.description,
            amount=main_amount,
            debit_amount=row.debit_amount or 0,
            credit_amount=row.credit_amount or 0,
            transaction_type=row.transaction_type,
            account_head_id=row.account_head_id if row.transaction_type == "DEBIT" else None,
            import_batch_id=batch.id,
            week_number=wn,
            week_year=wy,
            status="entered",
            entered_by=current_user.id,
            entered_at=datetime.utcnow(),
            entered_by_name=current_user.name if hasattr(current_user, 'name') else str(current_user.id),
        )
        db.add(txn)
        inserted += 1
        if row.transaction_type == "DEBIT":
            batch.debit_count += 1
        else:
            batch.credit_count += 1
        if row.balance is not None:
            last_balance = row.balance

    batch.row_count = inserted
    batch.duplicate_skipped = skipped

    # Update bank's last statement balance
    if last_balance is not None:
        acc.last_statement_balance = last_balance
    acc.last_import_at = datetime.utcnow()
    acc.updated_at = datetime.utcnow()

    db.commit()
    return {
        "inserted": inserted,
        "skipped_duplicates": skipped,
        "batch_id": batch.id,
        "message": f"Posted {inserted} transactions. {skipped} duplicates skipped."
    }


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# TRANSACTIONS (enhanced)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

@router.get("/transactions")
def list_transactions(
    account_id: Optional[int] = None,
    status: Optional[str] = None,
    transaction_type: Optional[str] = None,
    account_head_id: Optional[int] = None,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    week_number: Optional[int] = None,
    week_year: Optional[int] = None,
    import_batch_id: Optional[int] = None,
    limit: int = 200,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_login)
):
    q = db.query(BankTransaction)
    if account_id:
        q = q.filter(BankTransaction.account_id == account_id)
    if status:
        q = q.filter(BankTransaction.status == status)
    if transaction_type:
        q = q.filter(BankTransaction.transaction_type == transaction_type.upper())
    if account_head_id:
        q = q.filter(BankTransaction.account_head_id == account_head_id)
    if date_from:
        q = q.filter(BankTransaction.transaction_date >= date_from)
    if date_to:
        q = q.filter(BankTransaction.transaction_date <= date_to)
    if week_number is not None and week_year is not None:
        q = q.filter(BankTransaction.week_number == week_number, BankTransaction.week_year == week_year)
    if import_batch_id:
        q = q.filter(BankTransaction.import_batch_id == import_batch_id)
    txns = q.order_by(BankTransaction.transaction_date.desc(), BankTransaction.created_at.desc()).limit(limit).all()
    return [_ser_txn(t) for t in txns]


@router.post("/transactions")
def create_transaction(data: TransactionCreate, db: Session = Depends(get_db), current_user: User = Depends(require_login)):
    acc = db.query(BankAccount).filter(BankAccount.id == data.account_id, BankAccount.is_active == True).first()
    if not acc:
        raise HTTPException(404, "Bank account not found")
    txn_date = data.transaction_date or date.today()
    wn, wy, _, _ = get_week_info(txn_date)
    txn = BankTransaction(
        reference=_next_ref(db),
        account_id=data.account_id,
        transaction_date=txn_date,
        payee_name=data.payee_name,
        description=data.description,
        amount=data.amount,
        debit_amount=data.amount,
        credit_amount=0,
        transaction_type="DEBIT",
        account_head_id=data.account_head_id,
        week_number=wn,
        week_year=wy,
        status="entered",
        entered_by=current_user.id,
        entered_at=datetime.utcnow(),
        entered_by_name=current_user.name if hasattr(current_user, 'name') else str(current_user.id),
    )
    db.add(txn)
    db.commit()
    db.refresh(txn)
    return _ser_txn(txn)


@router.post("/transactions/{txn_id}/verify")
def verify_transaction(txn_id: int, data: TransactionActionRequest, db: Session = Depends(get_db), current_user: User = Depends(require_login)):
    txn = db.query(BankTransaction).filter(BankTransaction.id == txn_id).first()
    if not txn:
        raise HTTPException(404, "Transaction not found")
    if txn.status != "entered":
        raise HTTPException(400, f"Transaction is {txn.status}, cannot verify")
    txn.status = "verified"
    txn.verified_by = current_user.id
    txn.verified_at = datetime.utcnow()
    txn.verified_by_name = current_user.name if hasattr(current_user, 'name') else str(current_user.id)
    txn.verification_remarks = data.remarks
    txn.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(txn)
    return _ser_txn(txn)


@router.put("/transactions/{txn_id}/head")
def update_transaction_head(txn_id: int, account_head_id: Optional[int] = None, db: Session = Depends(get_db), current_user: User = Depends(require_login)):
    txn = db.query(BankTransaction).filter(BankTransaction.id == txn_id).first()
    if not txn:
        raise HTTPException(404, "Transaction not found")
    if txn.status == "approved":
        raise HTTPException(400, "Cannot modify approved transaction")
    txn.account_head_id = account_head_id
    txn.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(txn)
    return _ser_txn(txn)


@router.post("/transactions/{txn_id}/approve")
def approve_transaction(txn_id: int, data: TransactionActionRequest, db: Session = Depends(get_db), current_user: User = Depends(require_login)):
    txn = db.query(BankTransaction).filter(BankTransaction.id == txn_id).first()
    if not txn:
        raise HTTPException(404, "Transaction not found")
    if txn.status != "verified":
        raise HTTPException(400, "Transaction must be verified before approval")
    # Same-user rule: cannot approve what you verified
    if txn.verified_by == current_user.id:
        raise HTTPException(403, "You verified this transaction — a different user must approve it")
    txn.status = "approved"
    txn.approved_by = current_user.id
    txn.approved_at = datetime.utcnow()
    txn.approved_by_name = current_user.name if hasattr(current_user, 'name') else str(current_user.id)
    txn.approval_remarks = data.remarks
    txn.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(txn)
    return _ser_txn(txn)


@router.post("/transactions/{txn_id}/reject")
def reject_transaction(txn_id: int, data: TransactionActionRequest, db: Session = Depends(get_db), current_user: User = Depends(require_login)):
    txn = db.query(BankTransaction).filter(BankTransaction.id == txn_id).first()
    if not txn:
        raise HTTPException(404, "Transaction not found")
    if txn.status in ("approved", "rejected"):
        raise HTTPException(400, f"Transaction already {txn.status}")
    txn.status = "rejected"
    txn.rejection_remarks = data.remarks
    txn.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(txn)
    return _ser_txn(txn)


@router.delete("/transactions/{txn_id}")
def delete_transaction(txn_id: int, db: Session = Depends(get_db), current_user: User = Depends(require_admin)):
    txn = db.query(BankTransaction).filter(BankTransaction.id == txn_id).first()
    if not txn:
        raise HTTPException(404, "Transaction not found")
    if txn.status == "approved":
        raise HTTPException(400, "Cannot delete an approved transaction")
    db.delete(txn)
    db.commit()
    return {"message": "Transaction deleted"}


@router.post("/transactions/bulk-verify")
def bulk_verify(data: BulkActionRequest, db: Session = Depends(get_db), current_user: User = Depends(require_login)):
    updated = 0
    for txn_id in data.transaction_ids:
        txn = db.query(BankTransaction).filter(BankTransaction.id == txn_id).first()
        if txn and txn.status == "entered":
            txn.status = "verified"
            txn.verified_by = current_user.id
            txn.verified_at = datetime.utcnow()
            txn.verified_by_name = current_user.name if hasattr(current_user, 'name') else str(current_user.id)
            txn.verification_remarks = data.remarks
            txn.updated_at = datetime.utcnow()
            updated += 1
    db.commit()
    return {"updated": updated}


@router.post("/transactions/bulk-approve")
def bulk_approve(data: BulkActionRequest, db: Session = Depends(get_db), current_user: User = Depends(require_login)):
    updated = 0
    skipped_self_verify = 0
    for txn_id in data.transaction_ids:
        txn = db.query(BankTransaction).filter(BankTransaction.id == txn_id).first()
        if txn and txn.status == "verified":
            if txn.verified_by == current_user.id:
                skipped_self_verify += 1
                continue
            txn.status = "approved"
            txn.approved_by = current_user.id
            txn.approved_at = datetime.utcnow()
            txn.approved_by_name = current_user.name if hasattr(current_user, 'name') else str(current_user.id)
            txn.approval_remarks = data.remarks
            txn.updated_at = datetime.utcnow()
            updated += 1
    db.commit()
    return {"updated": updated, "skipped_self_verify": skipped_self_verify}


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# BANKING DASHBOARD
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

@router.get("/dashboard")
def dashboard(db: Session = Depends(get_db), current_user: User = Depends(require_login)):
    accounts = db.query(BankAccount).filter(BankAccount.is_active == True).all()
    today = date.today()
    yesterday = today - timedelta(days=1)

    # Week (Sun-Sat) boundaries for current week
    _, _, week_sun, week_sat = get_week_info(today)

    bank_cards = []
    total_balance = 0.0
    total_lien = 0.0
    total_actual = 0.0
    payments_this_week = 0.0

    for acc in accounts:
        bal = float(acc.last_statement_balance or acc.current_balance or 0)
        lien = float(acc.lien_amount or 0)
        actual = round(bal - lien, 2)
        total_balance += bal
        total_lien += lien
        total_actual += actual

        # Yesterday debits / credits
        yest_debits = db.query(func.sum(BankTransaction.debit_amount)).filter(
            BankTransaction.account_id == acc.id,
            BankTransaction.transaction_date == yesterday,
            BankTransaction.transaction_type == "DEBIT",
        ).scalar() or 0

        yest_credits = db.query(func.sum(BankTransaction.credit_amount)).filter(
            BankTransaction.account_id == acc.id,
            BankTransaction.transaction_date == yesterday,
            BankTransaction.transaction_type == "CREDIT",
        ).scalar() or 0

        # This week approved debits
        week_paid = db.query(func.sum(BankTransaction.debit_amount)).filter(
            BankTransaction.account_id == acc.id,
            BankTransaction.status == "approved",
            BankTransaction.transaction_type == "DEBIT",
            BankTransaction.transaction_date >= week_sun,
            BankTransaction.transaction_date <= week_sat,
        ).scalar() or 0
        payments_this_week += float(week_paid)

        bank_cards.append({
            **_ser_account(acc),
            "yesterday_debits": float(yest_debits),
            "yesterday_credits": float(yest_credits),
            "week_paid": float(week_paid),
        })

    return {
        "total_balance": round(total_balance, 2),
        "total_lien": round(total_lien, 2),
        "total_actual": round(total_actual, 2),
        "payments_this_week": round(payments_this_week, 2),
        "bank_cards": bank_cards,
        "week_start": week_sun.isoformat(),
        "week_end": week_sat.isoformat(),
        "today": today.isoformat(),
    }


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# WEEKLY PAYMENT BUCKETS
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

@router.get("/weekly-buckets")
def weekly_buckets(
    year: Optional[int] = None,
    month: Optional[int] = None,
    bank_id: Optional[int] = None,
    fy: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_login),
):
    today = date.today()
    target_year = year or today.year
    target_month = month or today.month
    target_fy = fy or current_fy()

    wks = weeks_in_month(target_year, target_month)
    num_weeks = len(wks)

    # Total monthly budget (sum across all active heads)
    budget_q = db.query(func.sum(BudgetMaster.planned_amount)).filter(
        BudgetMaster.financial_year == target_fy,
        BudgetMaster.month == target_month,
    )
    total_monthly_budget = float(budget_q.scalar() or 0)
    weekly_budget = round(total_monthly_budget / num_weeks, 2) if num_weeks else 0

    week_cards = []
    total_actual = 0.0

    for w in wks:
        # Approved DEBIT transactions in this Sun-Sat range
        q = db.query(func.sum(BankTransaction.debit_amount)).filter(
            BankTransaction.status == "approved",
            BankTransaction.transaction_type == "DEBIT",
            BankTransaction.transaction_date >= w["start"],
            BankTransaction.transaction_date <= w["end"],
        )
        if bank_id:
            q = q.filter(BankTransaction.account_id == bank_id)
        actual = float(q.scalar() or 0)
        total_actual += actual

        pct = round((actual / weekly_budget * 100), 1) if weekly_budget > 0 else 0
        if actual == 0:
            bucket_status = "Not started"
        elif actual >= weekly_budget * 1.0:
            bucket_status = "Over budget"
        elif actual >= weekly_budget * 0.8:
            bucket_status = "In progress"
        else:
            bucket_status = "Under budget"

        week_cards.append({
            "week_number": w["week_num"],
            "week_year": w["week_year"],
            "start": w["start"].isoformat(),
            "end": w["end"].isoformat(),
            "label": f"Week {w['week_num']} — {w['start'].strftime('%b %d')}–{w['end'].strftime('%b %d')}",
            "planned": weekly_budget,
            "actual": round(actual, 2),
            "pct": min(pct, 999),
            "status": bucket_status,
        })

    return {
        "year": target_year,
        "month": target_month,
        "month_name": date(target_year, target_month, 1).strftime("%B %Y"),
        "financial_year": target_fy,
        "total_monthly_budget": total_monthly_budget,
        "total_actual": round(total_actual, 2),
        "remaining": round(total_monthly_budget - total_actual, 2),
        "num_weeks": num_weeks,
        "weekly_budget": weekly_budget,
        "week_cards": week_cards,
    }


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# EXPENSE PIVOT REPORT
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def _pivot_data(fy: str, month: int, bank_id: Optional[int], db: Session) -> list:
    heads = db.query(AccountHead).filter(AccountHead.is_active == True).order_by(AccountHead.head_code).all()
    target_year = int(fy.split("-")[0])
    if month >= 4:
        target_year = int(fy.split("-")[0])
    else:
        target_year = int(fy.split("-")[0]) + 1
    first_day = date(target_year, month, 1)
    last_day = date(target_year, month, calendar.monthrange(target_year, month)[1])

    rows = []
    for h in heads:
        budget_row = db.query(BudgetMaster).filter(
            BudgetMaster.financial_year == fy,
            BudgetMaster.month == month,
            BudgetMaster.account_head_id == h.id,
        ).first()
        planned = float(budget_row.planned_amount) if budget_row else 0.0

        q = db.query(func.sum(BankTransaction.debit_amount)).filter(
            BankTransaction.account_head_id == h.id,
            BankTransaction.status == "approved",
            BankTransaction.transaction_type == "DEBIT",
            BankTransaction.transaction_date >= first_day,
            BankTransaction.transaction_date <= last_day,
        )
        if bank_id:
            q = q.filter(BankTransaction.account_id == bank_id)
        actual = float(q.scalar() or 0)
        variance = round(planned - actual, 2)
        rows.append({
            "head_code": h.head_code,
            "head_name": h.head_name,
            "category": h.category,
            "planned": planned,
            "actual": round(actual, 2),
            "variance": variance,
            "over_budget": variance < 0,
        })
    return rows


@router.get("/pivot")
def pivot_report(
    fy: Optional[str] = None,
    month: Optional[int] = None,
    bank_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_login),
):
    today = date.today()
    target_fy = fy or current_fy()
    target_month = month or today.month
    rows = _pivot_data(target_fy, target_month, bank_id, db)
    total_planned = sum(r["planned"] for r in rows)
    total_actual = sum(r["actual"] for r in rows)
    return {
        "financial_year": target_fy,
        "month": target_month,
        "month_name": date(today.year, target_month, 1).strftime("%B"),
        "rows": rows,
        "total_planned": round(total_planned, 2),
        "total_actual": round(total_actual, 2),
        "total_variance": round(total_planned - total_actual, 2),
    }


@router.get("/export/pivot.xlsx")
def export_pivot_xlsx(
    fy: Optional[str] = None,
    month: Optional[int] = None,
    bank_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_login),
):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(500, "openpyxl not installed")

    today = date.today()
    target_fy = fy or current_fy()
    target_month = month or today.month
    rows = _pivot_data(target_fy, target_month, bank_id, db)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Expense Pivot"

    # Header
    ws.append([f"Expense Pivot Report — {date(today.year, target_month, 1).strftime('%B')} {target_fy}"])
    ws.merge_cells("A1:F1")
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.append([])  # blank

    # Column headers
    headers = ["Head Code", "Head Name", "Category", "Budget (₹)", "Actual (₹)", "Variance (₹)"]
    ws.append(headers)
    header_row = ws.max_row
    for col, _ in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1a5902")
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    red_fill = PatternFill("solid", fgColor="FFCCCC")
    green_fill = PatternFill("solid", fgColor="CCFFCC")
    for r in rows:
        ws.append([r["head_code"], r["head_name"], r["category"] or "", r["planned"], r["actual"], r["variance"]])
        row_idx = ws.max_row
        fill = red_fill if r["over_budget"] else green_fill
        for col in range(4, 7):
            ws.cell(row=row_idx, column=col).fill = fill

    # Totals
    ws.append(["TOTAL", "", "",
               sum(r["planned"] for r in rows),
               sum(r["actual"] for r in rows),
               sum(r["variance"] for r in rows)])
    total_row = ws.max_row
    for col in range(1, 7):
        ws.cell(row=total_row, column=col).font = Font(bold=True)

    # Column widths
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 16

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    month_str = date(today.year, target_month, 1).strftime("%b_%Y")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=pivot_{month_str}.xlsx"}
    )


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# MANAGEMENT SUMMARY REPORT (Ravi Sir Report)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

@router.get("/report/management")
def management_report(db: Session = Depends(get_db), current_user: User = Depends(require_login)):
    today = date.today()
    fy = current_fy()
    target_month = today.month
    _, _, week_sun, week_sat = get_week_info(today)

    # ── Section A: Bank Position ─────────────────────────────────────────────
    accounts = db.query(BankAccount).filter(BankAccount.is_active == True).all()
    bank_position = []
    total_bal = 0.0
    total_lien = 0.0
    for acc in accounts:
        bal = float(acc.last_statement_balance or acc.current_balance or 0)
        lien = float(acc.lien_amount or 0)
        total_bal += bal
        total_lien += lien
        bank_position.append({
            "bank_name": acc.account_name,
            "account_number": acc.account_number,
            "balance": bal,
            "lien": lien,
            "actual": round(bal - lien, 2),
            "last_import_at": acc.last_import_at.isoformat() if acc.last_import_at else None,
        })

    # ── Section B: Month Budget vs Actual ────────────────────────────────────
    first_day = date(today.year, target_month, 1)
    last_day = date(today.year, target_month, calendar.monthrange(today.year, target_month)[1])

    total_budget = float(db.query(func.sum(BudgetMaster.planned_amount)).filter(
        BudgetMaster.financial_year == fy,
        BudgetMaster.month == target_month,
    ).scalar() or 0)

    total_spent = float(db.query(func.sum(BankTransaction.debit_amount)).filter(
        BankTransaction.status == "approved",
        BankTransaction.transaction_type == "DEBIT",
        BankTransaction.transaction_date >= first_day,
        BankTransaction.transaction_date <= last_day,
    ).scalar() or 0)

    # Over-budget heads
    pivot_rows = _pivot_data(fy, target_month, None, db)
    over_budget_heads = [r for r in pivot_rows if r["over_budget"]]

    # ── Section C: This Week ─────────────────────────────────────────────────
    week_planned_q = db.query(func.sum(BudgetMaster.planned_amount)).filter(
        BudgetMaster.financial_year == fy,
        BudgetMaster.month == target_month,
    ).scalar() or 0
    num_wks = len(weeks_in_month(today.year, today.month))
    week_planned = round(float(week_planned_q) / num_wks, 2) if num_wks else 0

    week_paid = float(db.query(func.sum(BankTransaction.debit_amount)).filter(
        BankTransaction.status == "approved",
        BankTransaction.transaction_type == "DEBIT",
        BankTransaction.transaction_date >= week_sun,
        BankTransaction.transaction_date <= week_sat,
    ).scalar() or 0)

    week_pending_amount = float(db.query(func.sum(BankTransaction.debit_amount)).filter(
        BankTransaction.status.in_(["entered", "verified"]),
        BankTransaction.transaction_type == "DEBIT",
        BankTransaction.transaction_date >= week_sun,
        BankTransaction.transaction_date <= week_sat,
    ).scalar() or 0)

    week_pending_count = db.query(func.count(BankTransaction.id)).filter(
        BankTransaction.status.in_(["entered", "verified"]),
        BankTransaction.transaction_date >= week_sun,
        BankTransaction.transaction_date <= week_sat,
    ).scalar() or 0

    # ── Section D: Pending Approvals Alert (older than 2 days) ───────────────
    two_days_ago = today - timedelta(days=2)
    old_pending = db.query(BankTransaction).filter(
        BankTransaction.status.in_(["entered", "verified"]),
        BankTransaction.transaction_date <= two_days_ago,
    ).order_by(BankTransaction.transaction_date.asc()).limit(50).all()

    pending_alerts = []
    for t in old_pending:
        acc = next((a for a in accounts if a.id == t.account_id), None)
        pending_alerts.append({
            "id": t.id,
            "reference": t.reference,
            "transaction_date": t.transaction_date.isoformat() if t.transaction_date else None,
            "description": t.description,
            "amount": float(t.debit_amount or t.amount or 0),
            "bank_name": acc.account_name if acc else "—",
            "status": t.status,
        })

    return {
        "generated_at": datetime.utcnow().isoformat(),
        "today": today.isoformat(),
        "financial_year": fy,
        "month_name": date(today.year, target_month, 1).strftime("%B %Y"),
        "week_range": f"{week_sun.strftime('%d %b')} – {week_sat.strftime('%d %b %Y')}",
        "section_a": {
            "bank_position": bank_position,
            "total_balance": round(total_bal, 2),
            "total_lien": round(total_lien, 2),
            "total_actual": round(total_bal - total_lien, 2),
        },
        "section_b": {
            "total_budget": total_budget,
            "total_spent": round(total_spent, 2),
            "balance_remaining": round(total_budget - total_spent, 2),
            "over_budget_heads": over_budget_heads,
        },
        "section_c": {
            "week_range": f"{week_sun.strftime('%d %b')} – {week_sat.strftime('%d %b %Y')}",
            "planned": week_planned,
            "paid_so_far": round(week_paid, 2),
            "pending_amount": round(week_pending_amount, 2),
            "pending_count": week_pending_count,
        },
        "section_d": {
            "pending_alerts": pending_alerts,
            "total_alerts": len(pending_alerts),
        },
    }


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# SEED (idempotent — run once on startup)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def seed_finance_data(db: Session):
    """Seed default banks + account heads if tables are empty."""
    # Default banks
    default_banks = [
        {"account_name": "KOTAK Bank", "bank_name": "Kotak Mahindra Bank", "branch_name": "Main Branch"},
        {"account_name": "BOB Bank", "bank_name": "Bank of Baroda", "branch_name": "Main Branch"},
        {"account_name": "HDFC Bank", "bank_name": "HDFC Bank", "branch_name": "Main Branch"},
    ]
    if db.query(func.count(BankAccount.id)).scalar() == 0:
        for b in default_banks:
            db.add(BankAccount(
                account_name=b["account_name"],
                bank_name=b["bank_name"],
                branch_name=b["branch_name"],
                is_active=True,
            ))
        db.commit()
        print("✅ Seeded default bank accounts (KOTAK, BOB, HDFC)")

    # Default account heads
    default_heads = [
        {"head_code": "NPD_PROD",  "head_name": "NPD & Production",           "category": "Production"},
        {"head_code": "COGS_WC",   "head_name": "COGS / Working Capital",     "category": "COGS"},
        {"head_code": "SALARY",    "head_name": "Salary Budget",               "category": "HR"},
        {"head_code": "AP_LOAN",   "head_name": "Accounts Payable / Loan / Equity", "category": "Finance"},
        {"head_code": "RENT_OPEX", "head_name": "Rent & OPEX",                "category": "Operations"},
    ]
    if db.query(func.count(AccountHead.id)).scalar() == 0:
        for h in default_heads:
            db.add(AccountHead(head_code=h["head_code"], head_name=h["head_name"], category=h["category"]))
        db.commit()
        print("✅ Seeded default account heads")
