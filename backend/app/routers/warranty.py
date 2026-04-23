from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import or_, func
from ..database import get_db
from ..models import Product, BOM, BOMComponent, ProductComponentSerial, Stage, Component, ProductCategory, TaxConfig
from ..auth import get_current_user
from typing import Optional, List
import datetime, io, openpyxl
from fastapi.responses import StreamingResponse

router = APIRouter()

def calculate_expiry(start_date: datetime.date, period: int, unit: str):
    if not start_date or not period: return None
    # Simple logic for months/years/days
    days = 0
    u = unit.lower()
    if 'year' in u: days = period * 365
    elif 'month' in u: days = period * 30
    else: days = period # default days
    return start_date + datetime.timedelta(days=days)

@router.post("/products/{id}/activate")
def activate_warranty(id: int, db: Session = Depends(get_db)):
    p = db.query(Product).filter(Product.id == id).first()
    if not p: raise HTTPException(404)
    
    now = datetime.date.today()
    p.warranty_start_date = now
    p.warranty_end_date = calculate_expiry(now, p.warranty_period, p.warranty_unit or "months")
    
    for cs in p.component_serials:
        cs.warranty_start_date = now
        cs.warranty_end_date = calculate_expiry(now, cs.warranty_period, cs.warranty_unit or "months")
        
    db.commit()
    db.refresh(p)
    return serialize_product(p)

def serialize_product(p: Product):
    try:
        return {
            "id": p.id,
            "title": p.title or "",
            "name": p.name or "",
            "serial_number": p.serial_number or "",
            "warranty_period": p.warranty_period or 0,
            "warranty_unit": p.warranty_unit or "months",
            "warranty_start_date": str(p.warranty_start_date) if p.warranty_start_date else None,
            "warranty_end_date": str(p.warranty_end_date) if p.warranty_end_date else None,
            "notes": p.notes or "",
            "custom_data": p.custom_data or {},
            "created_at": str(p.created_at) if p.created_at else "",
            "stage_name": p.stage.name if p.stage else None,
            "stage_color": p.stage.color if p.stage else None,
            "bom_id": p.bom_id,
            "bom_name": p.bom.name if p.bom else None,
            "component_serials": [{
                "bom_component_id": c.bom_component_id,
                "name": c.bom_component.name if c.bom_component else "",
                "serial_number": c.serial_number or "",
                "warranty_period": c.warranty_period,
                "warranty_unit": c.warranty_unit,
                "warranty_start_date": str(c.warranty_start_date) if c.warranty_start_date else None,
                "warranty_end_date": str(c.warranty_end_date) if c.warranty_end_date else None
            } for c in p.component_serials] if 'component_serials' in p.__dict__ else []
        }
    except Exception as e:
        print(f"Serialization error for product {getattr(p, 'id', 'unknown')}: {e}")
        return {"id": getattr(p, 'id', 0), "name": "Error loading product"}

@router.get("/products")
def get_products(
    search: Optional[str] = None, 
    stage_id: Optional[int] = None, 
    skip: int = 0, 
    limit: int = 50, 
    db: Session = Depends(get_db)
):
    try:
        # Efficiently get stage counts first (this doesn't need to join everything)
        stage_counts = {str(s_id): count for s_id, count in db.query(Product.stage_id, func.count(Product.id)).group_by(Product.stage_id).all() if s_id}

        q = db.query(Product).options(
            joinedload(Product.stage),
            joinedload(Product.bom)
        )
        
        if search:
            q = q.filter(or_(
                Product.title.ilike(f"%{search}%"),
                Product.name.ilike(f"%{search}%"),
                Product.serial_number.ilike(f"%{search}%")
            ))
        
        if stage_id:
            q = q.filter(Product.stage_id == stage_id)
            
        total = q.count()
        products = q.order_by(Product.id.desc()).offset(skip).limit(limit).all()
        
        return {
            "total": total, 
            "items": [serialize_product(p) for p in products],
            "stage_counts": stage_counts
        }
    except Exception as e:
        print(f"Error in get_products: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/products/{id}")
def get_product(id: int, db: Session = Depends(get_db)):
    p = db.query(Product).options(
        joinedload(Product.stage),
        joinedload(Product.bom),
        joinedload(Product.component_serials).joinedload(ProductComponentSerial.bom_component)
    ).filter(Product.id == id).first()
    if not p: raise HTTPException(404)
    return serialize_product(p)

@router.get("/products/{id}/navigation")
def get_product_navigation(id: int, db: Session = Depends(get_db)):
    prev_id = db.query(Product.id).filter(Product.id < id).order_by(Product.id.desc()).first()
    next_id = db.query(Product.id).filter(Product.id > id).order_by(Product.id.asc()).first()
    return {
        "prev": prev_id[0] if prev_id else None,
        "next": next_id[0] if next_id else None
    }

@router.post("/products")
def create_product(data: dict, db: Session = Depends(get_db)):
    try:
        comps = data.pop("component_serials", [])
        valid_fields = ["title", "name", "serial_number", "warranty_period", "warranty_unit", "stage_id", "notes", "custom_data", "bom_id"]
        product_data = {k: v for k, v in data.items() if k in valid_fields}
        
        # Convert empty serial_number to None to avoid unique constraint clash when empty
        if not product_data.get("serial_number"):
            product_data["serial_number"] = None

        # Default title to BOM name if title is empty
        if not product_data.get("title") and product_data.get("bom_id"):
            bom = db.query(BOM).filter(BOM.id == product_data["bom_id"]).first()
            if bom: product_data["title"] = bom.name

        p = Product(**product_data)
        db.add(p); db.commit(); db.refresh(p)

        for c in comps:
            # Remove helper fields like 'name' before saving to ProductComponentSerial
            cs_data = {k: v for k, v in c.items() if k in ["bom_component_id", "serial_number", "warranty_period", "warranty_unit"]}
            cs = ProductComponentSerial(**cs_data, product_id=p.id)
            db.add(cs)
        
        db.commit(); db.refresh(p)
        return serialize_product(p)
    except Exception as e:
        print(f"Error in create_product: {e}")
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))

@router.put("/products/{id}")
def update_product(id: int, data: dict, db: Session = Depends(get_db)):
    try:
        comps = data.pop("component_serials", [])
        p = db.query(Product).filter(Product.id == id).first()
        if not p: raise HTTPException(404)
        
        valid_fields = ["title", "name", "serial_number", "warranty_period", "warranty_unit", "stage_id", "notes", "custom_data", "bom_id"]
        for k, v in data.items():
            if k in valid_fields: setattr(p, k, v)
            
        # Convert empty serial_number to None to avoid unique constraint clash when empty
        if not p.serial_number:
            p.serial_number = None

        # Default title to BOM name if title is empty
        if not p.title and p.bom_id:
            bom = db.query(BOM).filter(BOM.id == p.bom_id).first()
            if bom: p.title = bom.name
        
        # Sync components
        db.query(ProductComponentSerial).filter(ProductComponentSerial.product_id == id).delete()
        for c in comps:
            cs_data = {k: v for k, v in c.items() if k in ["bom_component_id", "serial_number", "warranty_period", "warranty_unit"]}
            cs = ProductComponentSerial(**cs_data, product_id=p.id)
            db.add(cs)
        
        db.commit(); db.refresh(p)
        return serialize_product(p)
    except Exception as e:
        print(f"Error in update_product: {e}")
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))

def serialize_bom(b: BOM):
    return {
        "id": b.id,
        "name": b.name,
        "description": b.description,
        "warranty_period": b.warranty_period,
        "warranty_unit": b.warranty_unit,
        "created_at": str(b.created_at),
        "components": [{
            "id": c.id,
            "name": c.name,
            "part_number": c.part_number,
            "quantity": c.quantity,
            "warranty_period": c.warranty_period,
            "warranty_unit": c.warranty_unit,
            "sort_order": c.sort_order
        } for c in sorted(b.components, key=lambda x: x.sort_order)]
    }

@router.get("/boms")
def get_boms(db: Session = Depends(get_db)):
    boms = db.query(BOM).options(joinedload(BOM.components)).all()
    return [serialize_bom(b) for b in boms]

@router.get("/components")
def get_components(search: Optional[str] = None, db: Session = Depends(get_db)):
    q = db.query(Component)
    if search:
        q = q.filter(or_(
            Component.name.ilike(f"%{search}%"),
            Component.part_number.ilike(f"%{search}%"),
            Component.category.ilike(f"%{search}%")
        ))
    comps = q.all()
    return [{
        "id": c.id,
        "name": c.name,
        "category": c.category,
        "part_number": c.part_number,
        "product_type": c.product_type,
        "sales_price": c.sales_price,
        "sales_taxes": c.sales_taxes,
        "on_hand_qty": c.on_hand_qty,
        "image_url": c.image_url,
        "created_at": str(c.created_at)
    } for c in comps]

@router.post("/components")
def create_master_component(data: dict, db: Session = Depends(get_db)):
    try:
        valid_fields = ["name", "category", "part_number", "product_type", "sales_price", "sales_taxes", "on_hand_qty", "image_url"]
        comp_data = {k: v for k, v in data.items() if k in valid_fields}
        c = Component(**comp_data)
        db.add(c); db.commit(); db.refresh(c)
        return c
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))

@router.put("/components/{id}")
def update_master_component(id: int, data: dict, db: Session = Depends(get_db)):
    c = db.query(Component).filter(Component.id == id).first()
    if not c: raise HTTPException(404)
    valid_fields = ["name", "category", "part_number", "product_type", "sales_price", "sales_taxes", "on_hand_qty", "image_url"]
    for k, v in data.items():
        if k in valid_fields: setattr(c, k, v)
    db.commit(); db.refresh(c); return c

@router.delete("/components/{id}")
def delete_master_component(id: int, db: Session = Depends(get_db)):
    c = db.query(Component).filter(Component.id == id).first()
    if not c: raise HTTPException(404)
    db.delete(c); db.commit(); return {"message": "Deleted"}

@router.get("/components/export/excel")
def export_components(db: Session = Depends(get_db)):
    """Export all master components to Excel."""
    comps = db.query(Component).order_by(Component.name).all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Components"

    headers = ["ID", "Product Name", "Category", "Part Number", "Product Type",
               "Sales Price", "Sales Taxes (%)", "On Hand Qty", "Image URL"]
    ws.append(headers)

    # Style header row
    from openpyxl.styles import Font, PatternFill, Alignment
    header_fill = PatternFill("solid", fgColor="1a5402")
    for col, cell in enumerate(ws[1], 1):
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = [5,25,20,18,16,14,16,14,40][col-1]

    for c in comps:
        ws.append([
            c.id, c.name or "", c.category or "", c.part_number or "",
            c.product_type or "Storable", c.sales_price or 0,
            c.sales_taxes or 0, c.on_hand_qty or 0, c.image_url or ""
        ])

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=components.xlsx"})

@router.get("/components/template/excel")
def download_import_template():
    """Download blank import template."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Components"
    headers = ["Product Name", "Category", "Part Number", "Product Type",
               "Sales Price", "Sales Taxes (%)", "On Hand Qty", "Image URL"]
    ws.append(headers)

    from openpyxl.styles import Font, PatternFill, Alignment
    header_fill = PatternFill("solid", fgColor="1a5402")
    for col, cell in enumerate(ws[1], 1):
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = [25,20,18,16,14,16,14,40][col-1]

    # Add one example row
    ws.append(["Battery 60V", "Electronics", "BAT-60V-001", "Storable", 2500, 18, 50, ""])

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=components_template.xlsx"})

@router.post("/components/import/excel")
async def import_components(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """Bulk import components from Excel. Updates existing if Part Number matches, else creates."""
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(400, "Only .xlsx or .xls files are accepted")
    try:
        content = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active

        created = updated = skipped = 0
        errors = []

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row): continue  # skip empty rows
            try:
                name       = str(row[0]).strip() if row[0] else None
                category   = str(row[1]).strip() if row[1] else ""
                part_no    = str(row[2]).strip() if row[2] else None
                prod_type  = str(row[3]).strip() if row[3] else "Storable"
                price      = float(row[4]) if row[4] is not None else 0.0
                taxes      = float(row[5]) if row[5] is not None else 0.0
                qty        = float(row[6]) if row[6] is not None else 0.0
                image      = str(row[7]).strip() if row[7] else ""

                if not name:
                    errors.append(f"Row {row_idx}: Product Name is required — skipped"); skipped += 1; continue

                # Upsert by Part Number if provided
                existing = db.query(Component).filter(Component.part_number == part_no).first() if part_no else None
                if existing:
                    existing.name = name; existing.category = category
                    existing.product_type = prod_type; existing.sales_price = price
                    existing.sales_taxes = taxes; existing.on_hand_qty = qty
                    existing.image_url = image
                    updated += 1
                else:
                    c = Component(name=name, category=category, part_number=part_no,
                                  product_type=prod_type, sales_price=price,
                                  sales_taxes=taxes, on_hand_qty=qty, image_url=image)
                    db.add(c); created += 1
            except Exception as e:
                errors.append(f"Row {row_idx}: {str(e)}"); skipped += 1

        db.commit()
        return {
            "success": True,
            "created": created, "updated": updated, "skipped": skipped,
            "errors": errors,
            "message": f"Import complete: {created} created, {updated} updated, {skipped} skipped."
        }
    except Exception as e:
        db.rollback()
        raise HTTPException(500, f"Import failed: {str(e)}")

@router.get("/boms/{id}")
def get_bom(id: int, db: Session = Depends(get_db)):
    b = db.query(BOM).options(joinedload(BOM.components)).filter(BOM.id == id).first()
    if not b: raise HTTPException(404)
    return serialize_bom(b)

@router.post("/boms")
def create_bom(data: dict, db: Session = Depends(get_db)):
    comps = data.pop("components", [])
    b = BOM(**data)
    db.add(b); db.commit(); db.refresh(b)
    for c in comps:
        comp = BOMComponent(**c, bom_id=b.id)
        db.add(comp)
    db.commit(); db.refresh(b)
    return serialize_bom(b)

@router.put("/boms/{id}")
def update_bom(id: int, data: dict, db: Session = Depends(get_db)):
    try:
        comps = data.pop("components", [])
        b = db.query(BOM).filter(BOM.id == id).first()
        if not b: raise HTTPException(404)
        
        valid_fields = ["name", "description", "warranty_period", "warranty_unit"]
        for k, v in data.items():
            if k in valid_fields: setattr(b, k, v)
        
        # Smart sync for components
        existing_comp_ids = [c.id for c in b.components]
        incoming_comp_ids = [c.get("id") for c in comps if c.get("id")]
        
        # 1. Delete removed components (only if not used by others)
        for ec_id in existing_comp_ids:
            if ec_id not in incoming_comp_ids:
                db.query(BOMComponent).filter(BOMComponent.id == ec_id).delete()
        
        # 2. Update or Create
        for c_data in comps:
            cid = c_data.get("id")
            # Filter valid component fields
            cv_fields = ["name", "part_number", "quantity", "warranty_period", "warranty_unit", "sort_order"]
            cv_data = {k: v for k, v in c_data.items() if k in cv_fields}
            
            if cid and cid in existing_comp_ids:
                # Update existing
                db.query(BOMComponent).filter(BOMComponent.id == cid).update(cv_data)
            else:
                # Create new
                new_c = BOMComponent(**cv_data, bom_id=id)
                db.add(new_c)
        
        db.commit(); db.refresh(b)
        return serialize_bom(b)
    except Exception as e:
        print(f"Error in update_bom: {e}")
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))

# ─── Product Categories ─────────────────────────────────────────────────────

@router.get("/config/categories")
def get_categories(db: Session = Depends(get_db)):
    return [{"id": c.id, "name": c.name, "description": c.description or ""} for c in db.query(ProductCategory).order_by(ProductCategory.name).all()]

@router.post("/config/categories")
def create_category(data: dict, db: Session = Depends(get_db)):
    if not data.get("name"): raise HTTPException(400, "Name is required")
    try:
        c = ProductCategory(name=data["name"].strip(), description=data.get("description",""))
        db.add(c); db.commit(); db.refresh(c)
        return {"id": c.id, "name": c.name, "description": c.description or ""}
    except Exception as e:
        db.rollback(); raise HTTPException(500, str(e))

@router.put("/config/categories/{id}")
def update_category(id: int, data: dict, db: Session = Depends(get_db)):
    c = db.query(ProductCategory).filter(ProductCategory.id == id).first()
    if not c: raise HTTPException(404)
    if data.get("name"): c.name = data["name"].strip()
    c.description = data.get("description", c.description)
    db.commit(); db.refresh(c)
    return {"id": c.id, "name": c.name, "description": c.description or ""}

@router.delete("/config/categories/{id}")
def delete_category(id: int, db: Session = Depends(get_db)):
    c = db.query(ProductCategory).filter(ProductCategory.id == id).first()
    if not c: raise HTTPException(404)
    db.delete(c); db.commit(); return {"message": "Deleted"}

# ─── Tax Configurations ──────────────────────────────────────────────────────

@router.get("/config/taxes")
def get_taxes(db: Session = Depends(get_db)):
    return [{"id": t.id, "name": t.name, "rate": t.rate, "description": t.description or ""} for t in db.query(TaxConfig).order_by(TaxConfig.name).all()]

@router.post("/config/taxes")
def create_tax(data: dict, db: Session = Depends(get_db)):
    if not data.get("name"): raise HTTPException(400, "Name is required")
    try:
        t = TaxConfig(name=data["name"].strip(), rate=float(data.get("rate", 0)), description=data.get("description",""))
        db.add(t); db.commit(); db.refresh(t)
        return {"id": t.id, "name": t.name, "rate": t.rate, "description": t.description or ""}
    except Exception as e:
        db.rollback(); raise HTTPException(500, str(e))

@router.put("/config/taxes/{id}")
def update_tax(id: int, data: dict, db: Session = Depends(get_db)):
    t = db.query(TaxConfig).filter(TaxConfig.id == id).first()
    if not t: raise HTTPException(404)
    if data.get("name"): t.name = data["name"].strip()
    if "rate" in data: t.rate = float(data["rate"])
    t.description = data.get("description", t.description)
    db.commit(); db.refresh(t)
    return {"id": t.id, "name": t.name, "rate": t.rate, "description": t.description or ""}

@router.delete("/config/taxes/{id}")
def delete_tax(id: int, db: Session = Depends(get_db)):
    t = db.query(TaxConfig).filter(TaxConfig.id == id).first()
    if not t: raise HTTPException(404)
    db.delete(t); db.commit(); return {"message": "Deleted"}
