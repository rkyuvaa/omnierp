from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form, Response
from sqlalchemy.orm import Session
from pydantic import BaseModel
from typing import Optional, List
from datetime import date, datetime, timedelta
import os, shutil, uuid, math
from app.database import get_db
from app.models import User, Branch
from app.auth import get_current_user, require_admin
from app.hr_models import (
    HRAttendancePunch, HRAttendanceRecord, HREmployee,
    HRShift, HRHoliday, HRLeaveRequest, HROnDutyRequest,
    HRLeaveBalance, HRLeaveType, HRConfig
)

router = APIRouter()

# ── Day-abbreviation map ─────────────────────────────────────────────────────
DAY_MAP = {0: "Mon", 1: "Tue", 2: "Wed", 3: "Thu", 4: "Fri", 5: "Sat", 6: "Sun"}

# ── Status colour map (for frontend) ─────────────────────────────────────────
STATUS_COLORS = {
    "present":    "#22c55e",
    "late":       "#f59e0b",
    "absent":     "#ef4444",
    "half_day":   "#f97316",
    "leave":      "#6366f1",
    "on_duty":    "#06b6d4",
    "holiday":    "#8b5cf6",
    "weekly_off": "#94a3b8",
}

# ── Comp-Off Accrual Helper ───────────────────────────────────────────────────
def _process_comp_off(db: Session, record: HRAttendanceRecord):
    """After attendance consolidation, check if overtime qualifies for comp-off.
    Updates record.comp_off_hours and the employee's CO leave balance.
    Only runs if comp_off_enabled config is True and activation_date is set.
    """
    from datetime import date as _date
    # Read configs
    def _cfg(key, default=None):
        row = db.query(HRConfig).filter(HRConfig.key == key).first()
        return row.value if row else default

    enabled = _cfg("comp_off_enabled", False)
    if not enabled:
        # Clear any previously accrued comp_off_hours
        if record.comp_off_hours and record.comp_off_hours > 0:
            _reverse_comp_off(db, record)
            record.comp_off_hours = 0
        return

    # Only accrue from the activation date onward
    activation_str = _cfg("comp_off_activation_date", None)
    if activation_str:
        try:
            activation_date = _date.fromisoformat(str(activation_str))
            if record.date < activation_date:
                return
        except Exception:
            pass

    threshold = float(_cfg("comp_off_threshold_hours", 9.0))
    hours_per_day = float(_cfg("comp_off_hours_per_day", 8.0))
    co_type_id = _cfg("comp_off_leave_type_id", None)

    if not co_type_id:
        # Auto-find or create Comp-Off leave type
        co_type = db.query(HRLeaveType).filter(HRLeaveType.code == "CO").first()
        if not co_type:
            return  # Will be created by setup endpoint; skip silently
        co_type_id = co_type.id

    hours_worked = record.hours_worked or 0
    prev_comp_off_hours = record.comp_off_hours or 0

    if hours_worked > threshold and hours_per_day > 0:
        excess = hours_worked - threshold
        half_day_hours = hours_per_day / 2.0
        
        if excess >= hours_per_day:
            new_comp_off_hours = hours_per_day
        elif excess >= half_day_hours:
            new_comp_off_hours = half_day_hours
        else:
            new_comp_off_hours = 0.0
    else:
        new_comp_off_hours = 0.0

    delta_hours = new_comp_off_hours - prev_comp_off_hours
    if delta_hours == 0:
        return  # No change needed

    delta_days = round(delta_hours / hours_per_day, 4)
    record.comp_off_hours = new_comp_off_hours

    # Update leave balance
    year = record.date.year
    balance = db.query(HRLeaveBalance).filter(
        HRLeaveBalance.employee_id == record.employee_id,
        HRLeaveBalance.leave_type_id == co_type_id,
        HRLeaveBalance.year == year,
    ).first()
    if not balance:
        balance = HRLeaveBalance(
            employee_id=record.employee_id,
            leave_type_id=co_type_id,
            year=year,
            allocated_days=0,
            used_days=0,
            carry_forwarded=0,
            monthly_limit=0,
        )
        db.add(balance)
        db.flush()

    balance.allocated_days = round(max(0.0, balance.allocated_days + delta_days), 4)


def _reverse_comp_off(db: Session, record: HRAttendanceRecord):
    """Reverse previously accrued comp-off hours for this record."""
    from datetime import date as _date
    if not record.comp_off_hours or record.comp_off_hours <= 0:
        return

    def _cfg(key, default=None):
        row = db.query(HRConfig).filter(HRConfig.key == key).first()
        return row.value if row else default

    hours_per_day = float(_cfg("comp_off_hours_per_day", 8.0))
    co_type_id = _cfg("comp_off_leave_type_id", None)
    if not co_type_id:
        co_type = db.query(HRLeaveType).filter(HRLeaveType.code == "CO").first()
        if not co_type:
            return
        co_type_id = co_type.id

    delta_days = round(record.comp_off_hours / hours_per_day, 4)
    year = record.date.year
    balance = db.query(HRLeaveBalance).filter(
        HRLeaveBalance.employee_id == record.employee_id,
        HRLeaveBalance.leave_type_id == co_type_id,
        HRLeaveBalance.year == year,
    ).first()
    if balance:
        balance.allocated_days = round(max(0.0, balance.allocated_days - delta_days), 4)

# ── Haversine Proximity Helpers ───────────────────────────────────────────────
def haversine_distance(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    """Calculate the great-circle distance between two GPS points in meters."""
    R = 6371000.0  # Earth's radius in meters
    phi1 = math.radians(lat1)
    phi2 = math.radians(lat2)
    delta_phi = math.radians(lat2 - lat1)
    delta_lambda = math.radians(lon2 - lon1)
    
    a = (math.sin(delta_phi / 2.0) ** 2 +
         math.cos(phi1) * math.cos(phi2) *
         math.sin(delta_lambda / 2.0) ** 2)
         
    c = 2.0 * math.atan2(math.sqrt(a), math.sqrt(1.0 - a))
    return R * c

def resolve_punch_location(p, active_branches) -> str:
    """If mobile punch falls in a branch radius, returns branch name, else original location_name."""
    if p.source == "mobile" and p.latitude is not None and p.longitude is not None:
        for b in active_branches:
            if b.latitude is not None and b.longitude is not None:
                dist = haversine_distance(p.latitude, p.longitude, b.latitude, b.longitude)
                radius = b.radius if b.radius is not None else 100.0
                if dist <= radius:
                    return b.name
    return p.location_name

class PunchManual(BaseModel):
    employee_id: int
    punch_time: datetime
    source: str = "manual"

class BulkPunch(BaseModel):
    biometric_id: str
    punch_time: datetime
    raw_uid: str

class BulkSync(BaseModel):
    machine_id: int
    punches: List[BulkPunch]

class AttendanceCorrect(BaseModel):
    employee_id: int
    date: date
    status: str
    check_in: Optional[datetime] = None
    check_out: Optional[datetime] = None
    correction_reason: str

class RecalculateRequest(BaseModel):
    employee_id: int
    date: date

def _is_sandwich_day(db: Session, employee_id: int, target_date: date) -> bool:
    # preceding date & following date
    prev_date = target_date - timedelta(days=1)
    next_date = target_date + timedelta(days=1)
    
    rec_prev = db.query(HRAttendanceRecord).filter(
        HRAttendanceRecord.employee_id == employee_id,
        HRAttendanceRecord.date == prev_date
    ).first()
    rec_next = db.query(HRAttendanceRecord).filter(
        HRAttendanceRecord.employee_id == employee_id,
        HRAttendanceRecord.date == next_date
    ).first()
    
    def is_absent_or_leave(rec, dt):
        if rec:
            return rec.status in ["absent", "leave", "sandwich_lop"]
        today = date.today()
        if dt < today:
            # Past working days without records default to absent
            return True
        return False
        
    return is_absent_or_leave(rec_prev, prev_date) and is_absent_or_leave(rec_next, next_date)

# ── Core: compute attendance status for one employee on one date ──────────────
def compute_record(db: Session, employee_id: int, target_date: date):
    emp = db.query(HREmployee).filter(HREmployee.id == employee_id).first()
    if not emp:
        return None

    # Get or create record
    record = db.query(HRAttendanceRecord).filter(
        HRAttendanceRecord.employee_id == employee_id,
        HRAttendanceRecord.date == target_date
    ).first()
    
    if record and record.corrected_by is not None:
        return record
        
    if not record:
        record = HRAttendanceRecord(employee_id=employee_id, date=target_date)
        db.add(record)

    # 1. Check for physical punches first
    day_start = datetime.combine(target_date, datetime.min.time())
    day_end = datetime.combine(target_date, datetime.max.time())
    punches = db.query(HRAttendancePunch).filter(
        HRAttendancePunch.employee_id == employee_id,
        HRAttendancePunch.punch_time >= day_start,
        HRAttendancePunch.punch_time <= day_end
    ).order_by(HRAttendancePunch.punch_time).all()

    if punches:
        check_in = punches[0].punch_time
        check_out = None
        if len(punches) > 1:
            # Avoid treating accidental double-punches (within 5 minutes) as check-out
            if (punches[-1].punch_time - check_in).total_seconds() >= 300:
                check_out = punches[-1].punch_time
                
        record.check_in = check_in
        record.check_out = check_out
        record.check_in_photo = punches[0].photo_url
        record.check_in_location = punches[0].location_name

        hours_worked = 0
        if check_out and check_out != check_in:
            hours_worked = (check_out - check_in).total_seconds() / 3600
        record.hours_worked = round(hours_worked, 2)

        # Clear active leave/on-duty references since the employee worked physically
        record.leave_request_id = None
        record.onduty_request_id = None

        shift = emp.shift
        # Compare against shift
        if shift:
            shift_start_h, shift_start_m = map(int, shift.start_time.split(":"))
            shift_end_h, shift_end_m = map(int, shift.end_time.split(":"))
            shift_start_dt = datetime.combine(target_date, datetime.min.time().replace(hour=shift_start_h, minute=shift_start_m))
            shift_end_dt = datetime.combine(target_date, datetime.min.time().replace(hour=shift_end_h, minute=shift_end_m))
            shift_duration = (shift_end_dt - shift_start_dt).total_seconds() / 3600

            grace_dt = shift_start_dt + timedelta(minutes=shift.grace_minutes)
            is_late = check_in > grace_dt
            late_minutes = max(0, int((check_in - grace_dt).total_seconds() / 60)) if is_late else 0

            record.is_late = is_late
            record.late_minutes = late_minutes

            # Half day thresholds
            half_day_late_dt = shift_start_dt + timedelta(minutes=shift.half_day_late_minutes or 240)
            half_day_early_dt = shift_end_dt - timedelta(minutes=shift.half_day_early_minutes or 240)

            if check_out and hours_worked < shift.half_day_hours:
                record.status = "half_day"
            elif check_in > half_day_late_dt:
                record.status = "half_day"
            elif check_out and check_out < half_day_early_dt:
                record.status = "half_day"
            elif is_late:
                record.status = "late"
            else:
                record.status = "present"

            # Early departure
            if check_out and check_out < shift_end_dt:
                record.left_early = True
                record.early_by_minutes = int((shift_end_dt - check_out).total_seconds() / 60)
        else:
            # Default logic if no shift is assigned
            if hours_worked > 0 and hours_worked < 4.0:
                record.status = "half_day"
            else:
                record.status = "present"

        db.commit()
        # Process comp-off accrual for this attendance record
        _process_comp_off(db, record)
        db.commit()
        return record

    # 2. If NO punches exist, fall back to passive statuses in priority order
    # Check holiday
    holiday = db.query(HRHoliday).filter(
        HRHoliday.date == target_date,
        HRHoliday.is_active == True,
        (HRHoliday.branch_id == emp.branch_id) | (HRHoliday.branch_id == None)
    ).first()

    if holiday:
        record.status = "holiday"
        record.check_in = None
        record.check_out = None
        record.hours_worked = 0
        db.commit()
        return record

    # Approved leave?
    leave_req = db.query(HRLeaveRequest).filter(
        HRLeaveRequest.employee_id == employee_id,
        HRLeaveRequest.from_date <= target_date,
        HRLeaveRequest.to_date >= target_date,
        HRLeaveRequest.status.in_(["approved", "auto_approved"])
    ).first()
    if leave_req:
        record.status = "leave"
        record.leave_request_id = leave_req.id
        record.check_in = None
        record.check_out = None
        record.hours_worked = 0
        db.commit()
        return record

    # Approved on-duty?
    od_req = db.query(HROnDutyRequest).filter(
        HROnDutyRequest.employee_id == employee_id,
        HROnDutyRequest.date == target_date,
        HROnDutyRequest.status.in_(["approved", "auto_approved"])
    ).first()
    if od_req:
        record.status = "on_duty"
        record.onduty_request_id = od_req.id
        record.check_in = None
        record.check_out = None
        record.hours_worked = 0
        db.commit()
        return record

    # Check weekly off
    shift = emp.shift
    from app.routers.hr_config import get_hr_config
    global_working_days = get_hr_config(db, "working_days", ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat"])
    
    is_wo = False
    if shift and shift.working_days:
        day_abbr = DAY_MAP[target_date.weekday()]
        if day_abbr not in shift.working_days:
            is_wo = True
    else:
        day_abbr = DAY_MAP[target_date.weekday()]
        if day_abbr not in global_working_days:
            is_wo = True

    if is_wo:
        enable_sandwich = get_hr_config(db, "enable_sandwich_highlight", True)
        auto_deduct = get_hr_config(db, "auto_deduct_sandwich", False)
        
        is_sandwich = _is_sandwich_day(db, employee_id, target_date)
        
        # Holiday Check: if Sunday is marked as holiday, sandwich doesn't apply
        holiday = db.query(HRHoliday).filter(
            HRHoliday.date == target_date,
            HRHoliday.is_active == True,
            (HRHoliday.branch_id == emp.branch_id) | (HRHoliday.branch_id == None)
        ).first()
        
        if not holiday and enable_sandwich and is_sandwich:
            if auto_deduct:
                # Automatically deduct unless HR manually ignored it
                if not record.correction_reason or "Ignored" not in (record.correction_reason or ""):
                    record.status = "sandwich_lop"
            else:
                # Sandwich detected but auto-deduct off. Keep as weekly_off (revert from sandwich_lop if it was one)
                if record.status == "sandwich_lop":
                    record.status = "weekly_off"
        else:
            # Not a sandwich day, revert status to weekly_off if it was sandwich_lop
            if record.status == "sandwich_lop":
                record.status = "weekly_off"
        
        if record.status != "sandwich_lop":
            record.status = "weekly_off"
            
        record.check_in = None
        record.check_out = None
        record.hours_worked = 0
        db.commit()
        return record

    # Reaching here means no punches and not a holiday/leave/on-duty/weekly-off -> absent
    record.status = "absent"
    record.check_in = None
    record.check_out = None
    record.hours_worked = 0
    db.commit()
    return record

# ── Routes ────────────────────────────────────────────────────────────────────
@router.post("/punch/mobile")
async def mobile_punch(
    employee_id: int = Form(...),
    latitude: float = Form(...),
    longitude: float = Form(...),
    location_name: str = Form(...),
    photo: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    if not current_user.is_superadmin:
        from app.auth import get_current_employee
        emp_resolved = get_current_employee(current_user, db)
        if employee_id != emp_resolved.id:
            raise HTTPException(status_code=403, detail="Access denied. You can only punch for yourself.")

    emp = db.query(HREmployee).filter(HREmployee.id == employee_id).first()
    if not emp: raise HTTPException(404, "Employee not found")

    # Save photo
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    upload_dir = os.path.join(BASE_DIR, "static", "uploads", "attendance")
    os.makedirs(upload_dir, exist_ok=True)
    ext = os.path.splitext(photo.filename)[1] or ".jpg"
    filename = f"punch_{employee_id}_{uuid.uuid4().hex}{ext}"
    filepath = os.path.join(upload_dir, filename)
    with open(filepath, "wb") as f:
        shutil.copyfileobj(photo.file, f)

    # Proximity matching with branches
    active_branches = db.query(Branch).filter(Branch.is_active == True).all()
    resolved_location_name = location_name
    if latitude is not None and longitude is not None:
        for b in active_branches:
            if b.latitude is not None and b.longitude is not None:
                dist = haversine_distance(latitude, longitude, b.latitude, b.longitude)
                radius = b.radius if b.radius is not None else 100.0
                if dist <= radius:
                    resolved_location_name = b.name
                    break

    punch_time = datetime.now()
    punch = HRAttendancePunch(
        employee_id=employee_id,
        punch_time=punch_time,
        source="mobile",
        photo_url=f"/api/uploads/attendance/{filename}",
        latitude=latitude,
        longitude=longitude,
        location_name=resolved_location_name,
    )
    db.add(punch); db.commit()

    # Recompute daily record
    compute_record(db, employee_id, punch_time.date())
    return {"message": "Punch recorded", "punch_time": str(punch_time), "location": resolved_location_name}

@router.post("/punch/manual")
def manual_punch(data: PunchManual, db: Session = Depends(get_db), current_user: User = Depends(require_admin)):
    punch = HRAttendancePunch(
        employee_id=data.employee_id,
        punch_time=data.punch_time,
        source="manual",
    )
    db.add(punch); db.commit()
    compute_record(db, data.employee_id, data.punch_time.date())
    return {"message": "Manual punch recorded"}

@router.get("/my-today")
def my_today_status(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    from app.auth import get_current_employee_optional
    from app.hr_models import HREmployee
    from sqlalchemy import func

    emp = get_current_employee_optional(current_user, db)
    
    # Robust fallback search if not linked yet
    if not emp and current_user.email:
        emp = db.query(HREmployee).filter(func.lower(HREmployee.email) == current_user.email.lower()).first()
        if emp:
            emp.user_id = current_user.id
            db.commit()
            db.refresh(emp)

    # Auto-create Employee profile for Superadmin/Admin if missing
    if not emp and current_user.is_superadmin:
        try:
            emp = HREmployee(
                user_id=current_user.id,
                employee_id=f"ADM{current_user.id:03d}",
                name=current_user.name or "Admin",
                email=current_user.email,
                joining_date=datetime.now().date(),
                status="active"
            )
            db.add(emp)
            db.commit()
            db.refresh(emp)
        except Exception as e:
            db.rollback()
            print(f"Error auto-creating superadmin employee: {e}")

    if not emp:
        return {
            "status": "no_employee",
            "employee_id": None,
            "enable_mobile_punch": False,
            "check_in": None,
            "check_out": None,
        }

    today = datetime.now().date()
    record = compute_record(db, emp.id, today)
    if not record:
        return {
            "status": "no_record",
            "employee_id": emp.id,
            "employee_name": emp.name,
            "enable_mobile_punch": emp.enable_mobile_punch or False,
            "check_in": None,
            "check_out": None,
            "hours_worked": 0,
        }
    return {
        "employee_id": emp.id,
        "employee_name": emp.name,
        "enable_mobile_punch": emp.enable_mobile_punch or False,
        "date": str(today),
        "status": record.status,
        "check_in": str(record.check_in) if record.check_in else None,
        "check_out": str(record.check_out) if record.check_out else None,
        "hours_worked": record.hours_worked,
        "is_late": record.is_late,
        "late_minutes": record.late_minutes,
        "check_in_photo": record.check_in_photo,
        "check_in_location": record.check_in_location,
        "status_color": STATUS_COLORS.get(record.status, "#94a3b8"),
    }

@router.get("/today/{employee_id}")
def today_status(employee_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    from app.auth import is_hr_admin
    if not is_hr_admin(current_user, db):
        from app.auth import get_current_employee_optional
        emp_resolved = get_current_employee_optional(current_user, db)
        if not emp_resolved or employee_id != emp_resolved.id:
            return {"status": "no_record"}

    today = datetime.now().date()
    record = compute_record(db, employee_id, today)
    if not record:
        return {"status": "no_record"}
    return {
        "employee_id": employee_id,
        "date": str(today),
        "status": record.status,
        "check_in": str(record.check_in) if record.check_in else None,
        "check_out": str(record.check_out) if record.check_out else None,
        "hours_worked": record.hours_worked,
        "is_late": record.is_late,
        "late_minutes": record.late_minutes,
        "check_in_photo": record.check_in_photo,
        "check_in_location": record.check_in_location,
        "status_color": STATUS_COLORS.get(record.status, "#94a3b8"),
    }

@router.get("/records")
def get_records(
    month: int,
    year: int,
    employee_id: Optional[int] = None,
    branch_id: Optional[int] = None,
    department_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    from app.auth import is_hr_admin
    if not is_hr_admin(current_user, db):
        from app.auth import get_current_employee_optional
        emp_resolved = get_current_employee_optional(current_user, db)
        if not emp_resolved:
            return {}
        if employee_id and employee_id != emp_resolved.id:
            raise HTTPException(status_code=403, detail="Access denied. You can only view your own records.")
        employee_id = emp_resolved.id
        branch_id = None
        department_id = None

    from sqlalchemy import extract
    q = db.query(HRAttendanceRecord).filter(
        extract('month', HRAttendanceRecord.date) == month,
        extract('year', HRAttendanceRecord.date) == year,
    )
    if employee_id:
        q = q.filter(HRAttendanceRecord.employee_id == employee_id)

    records = q.all()
    result = {}
    for r in records:
        is_paid = True
        if r.status == "leave" and r.leave_request:
            is_paid = r.leave_request.leave_type.is_paid if r.leave_request.leave_type else False

        key = str(r.employee_id)
        if key not in result:
            result[key] = {}
        result[key][str(r.date)] = {
            "status": r.status,
            "is_paid": is_paid,
            "color": STATUS_COLORS.get(r.status, "#94a3b8"),
            "check_in": str(r.check_in) if r.check_in else None,
            "check_out": str(r.check_out) if r.check_out else None,
            "hours_worked": r.hours_worked,
            "is_late": r.is_late,
            "late_minutes": r.late_minutes,
            "corrected_by": r.corrected_by,
        }
    return result

@router.post("/correct")
def correct_attendance(data: AttendanceCorrect, db: Session = Depends(get_db), current_user: User = Depends(require_admin)):
    record = db.query(HRAttendanceRecord).filter(
        HRAttendanceRecord.employee_id == data.employee_id,
        HRAttendanceRecord.date == data.date
    ).first()
    if not record:
        record = HRAttendanceRecord(employee_id=data.employee_id, date=data.date)
        db.add(record)
    record.status = data.status
    record.check_in = data.check_in
    record.check_out = data.check_out
    record.correction_reason = data.correction_reason
    record.corrected_by = current_user.id
    if data.check_in and data.check_out:
        record.hours_worked = round((data.check_out - data.check_in).total_seconds() / 3600, 2)
    db.commit()
    return {"message": "Attendance corrected"}

@router.post("/recalculate-day")
def recalculate_day(data: RecalculateRequest, db: Session = Depends(get_db), current_user: User = Depends(require_admin)):
    record = db.query(HRAttendanceRecord).filter(
        HRAttendanceRecord.employee_id == data.employee_id,
        HRAttendanceRecord.date == data.date
    ).first()
    if record:
        record.corrected_by = None
        record.correction_reason = None
        record.check_in = None
        record.check_out = None
        record.hours_worked = 0
        db.commit()
    compute_record(db, data.employee_id, data.date)
    return {"message": "Recalculated successfully"}

@router.get("/punches/{employee_id}")
def get_punches(employee_id: int, target_date: date = None,
                db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    from app.auth import is_hr_admin
    if not is_hr_admin(current_user, db):
        from app.auth import get_current_employee_optional
        emp_resolved = get_current_employee_optional(current_user, db)
        if not emp_resolved or employee_id != emp_resolved.id:
            return []

    q = db.query(HRAttendancePunch).filter(HRAttendancePunch.employee_id == employee_id)
    if target_date:
        day_start = datetime.combine(target_date, datetime.min.time())
        day_end = datetime.combine(target_date, datetime.max.time())
        q = q.filter(HRAttendancePunch.punch_time.between(day_start, day_end))
    punches = q.order_by(HRAttendancePunch.punch_time).all()
    active_branches = db.query(Branch).filter(Branch.is_active == True).all()
    return [{
        "id": p.id, "punch_time": str(p.punch_time),
        "source": p.source, "photo_url": p.photo_url,
        "latitude": p.latitude, "longitude": p.longitude,
        "location_name": resolve_punch_location(p, active_branches),
    } for p in punches]

@router.get("/punches")
def get_all_punches(
    month: int,
    year: int,
    employee_id: Optional[int] = None,
    branch_id: Optional[int] = None,
    department_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    from app.auth import is_hr_admin
    if not is_hr_admin(current_user, db):
        from app.auth import get_current_employee_optional
        emp_resolved = get_current_employee_optional(current_user, db)
        if not emp_resolved:
            return []
        employee_id = emp_resolved.id
        branch_id = None
        department_id = None

    from sqlalchemy import extract
    q = db.query(HRAttendancePunch).join(HREmployee, HRAttendancePunch.employee_id == HREmployee.id)
    
    q = q.filter(
        extract('month', HRAttendancePunch.punch_time) == month,
        extract('year', HRAttendancePunch.punch_time) == year
    )
    
    if employee_id:
        q = q.filter(HRAttendancePunch.employee_id == employee_id)
    if branch_id:
        q = q.filter(HREmployee.branch_id == branch_id)
    if department_id:
        q = q.filter(HREmployee.department_id == department_id)

    active_branches = db.query(Branch).filter(Branch.is_active == True).all()
    punches = q.order_by(HRAttendancePunch.punch_time.desc()).all()
    
    return [{
        "id": p.id,
        "employee_id": p.employee_id,
        "employee_name": p.employee.name if p.employee else "-",
        "employee_code": p.employee.employee_id if p.employee else "-",
        "punch_time": str(p.punch_time),
        "source": p.source,
        "photo_url": p.photo_url,
        "latitude": p.latitude,
        "longitude": p.longitude,
        "location_name": resolve_punch_location(p, active_branches),
    } for p in punches]

class RecomputeRequest(BaseModel):
    month: int
    year: int

@router.post("/recompute")
def recompute_month(data: RecomputeRequest, db: Session = Depends(get_db), current_user: User = Depends(require_admin)):
    # Get all active employees
    employees = db.query(HREmployee).filter(HREmployee.is_active == True).all()
    
    import calendar
    _, last_day = calendar.monthrange(data.year, data.month)
    
    for emp in employees:
        for day in range(1, last_day + 1):
            target_date = date(data.year, data.month, day)
            # Only recompute if it's not in the future
            if target_date > date.today():
                continue
            compute_record(db, emp.id, target_date)
            
    return {"message": "Recomputation complete"}

@router.post("/sync/bulk")
def bulk_sync(data: BulkSync, db: Session = Depends(get_db)):
    """Bulk upload punches from a local sync agent"""
    imported = 0
    skipped = 0
    recompute_dates = set() # (employee_id, date)
    
    for p in data.punches:
        # Find employee by biometric_id
        emp = db.query(HREmployee).filter(HREmployee.biometric_id == p.biometric_id).first()
        if not emp:
            skipped += 1
            continue
            
        # Check if punch exists
        exists = db.query(HRAttendancePunch).filter(
            HRAttendancePunch.employee_id == emp.id,
            HRAttendancePunch.punch_time == p.punch_time
        ).first()
        
        if exists:
            skipped += 1
            continue
            
        punch = HRAttendancePunch(
            employee_id=emp.id,
            punch_time=p.punch_time,
            source="biometric",
            machine_id=data.machine_id,
            raw_punch_uid=p.raw_uid
        )
        db.add(punch)
        recompute_dates.add((emp.id, p.punch_time.date()))
        imported += 1
        
    db.commit()
    
    # Recompute daily records for all affected employees and dates
    for emp_id, target_date in recompute_dates:
        compute_record(db, emp_id, target_date)
        
    return {"message": "Bulk sync complete", "imported": imported, "skipped": skipped}


# ── Sandwich Leave Endpoints ───────────────────────────────────────────────
class SandwichDecision(BaseModel):
    employee_id: int
    date: date
    action: str  # deduct or ignore
    reason: Optional[str] = None

@router.get("/sandwich-leaves")
def get_sandwich_leaves(
    month: int,
    year: int,
    branch_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin)
):
    import calendar
    _, last_day = calendar.monthrange(year, month)
    sundays = []
    for d in range(1, last_day + 1):
        dt = date(year, month, d)
        if dt.weekday() == 6: # Sunday
            sundays.append(dt)
            
    q = db.query(HREmployee).filter(HREmployee.is_active == True)
    if branch_id:
        q = q.filter(HREmployee.branch_id == branch_id)
    employees = q.all()
    
    sandwich_list = []
    for emp in employees:
        for sun in sundays:
            sat = sun - timedelta(days=1)
            mon = sun + timedelta(days=1)
            
            rec_sun = db.query(HRAttendanceRecord).filter(
                HRAttendanceRecord.employee_id == emp.id,
                HRAttendanceRecord.date == sun
            ).first()
            
            rec_sat = db.query(HRAttendanceRecord).filter(
                HRAttendanceRecord.employee_id == emp.id,
                HRAttendanceRecord.date == sat
            ).first()
            rec_mon = db.query(HRAttendanceRecord).filter(
                HRAttendanceRecord.employee_id == emp.id,
                HRAttendanceRecord.date == mon
            ).first()
            
            sat_status = rec_sat.status if rec_sat else ("absent" if sat < date.today() else "no_record")
            mon_status = rec_mon.status if rec_mon else ("absent" if mon < date.today() else "no_record")
            
            is_pot = sat_status in ["absent", "leave", "sandwich_lop"] and mon_status in ["absent", "leave", "sandwich_lop"]
            
            if is_pot:
                sandwich_list.append({
                    "employee_id": emp.id,
                    "employee_name": emp.name,
                    "employee_code": emp.employee_id,
                    "date": str(sun),
                    "sat_status": sat_status,
                    "mon_status": mon_status,
                    "current_status": rec_sun.status if rec_sun else "weekly_off",
                    "reason": rec_sun.correction_reason if rec_sun else None,
                    "decided_by": rec_sun.corrected_by if rec_sun else None,
                })
                
    return sandwich_list


@router.post("/sandwich-decision")
def post_sandwich_decision(
    data: SandwichDecision,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin)
):
    record = db.query(HRAttendanceRecord).filter(
        HRAttendanceRecord.employee_id == data.employee_id,
        HRAttendanceRecord.date == data.date
    ).first()
    
    if not record:
        record = HRAttendanceRecord(
            employee_id=data.employee_id,
            date=data.date,
            status="weekly_off"
        )
        db.add(record)
        
    if data.action == "deduct":
        record.status = "sandwich_lop"
        record.correction_reason = data.reason or "Deducted as Sandwich LOP"
        record.corrected_by = current_user.id
    elif data.action == "ignore":
        record.status = "weekly_off"
        record.correction_reason = data.reason or "Ignored sandwich LOP"
        record.corrected_by = current_user.id
        
    db.commit()
    return {"message": f"Deduction {data.action}ed successfully."}


@router.get("/import/template")
def get_attendance_template(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    import openpyxl
    from openpyxl.worksheet.datavalidation import DataValidation
    import io

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance Import Template"
    headers = ["employee_id", "employee_name", "date", "status", "check_in", "check_out"]
    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=h)

    # Add standard status dropdown list validation in column D
    dv = DataValidation(
        type="list", 
        formula1='"present,absent,half_day,weekly_off,holiday,leave,on_duty,late"', 
        allow_blank=True
    )
    dv.prompt = "Select status"
    dv.promptTitle = "Allowed Statuses"
    dv.error = "Must be present, absent, half_day, weekly_off, holiday, leave, on_duty, or late"
    dv.errorTitle = "Invalid Status"
    ws.add_data_validation(dv)
    dv.add("D2:D1000")

    # Pre-populate active employees
    employees = db.query(HREmployee).filter(HREmployee.is_active == True).order_by(HREmployee.employee_id).all()
    for row_idx, emp in enumerate(employees, 2):
        ws.cell(row=row_idx, column=1, value=emp.employee_id)
        ws.cell(row=row_idx, column=2, value=emp.name)
        ws.cell(row=row_idx, column=3, value=str(date.today() - timedelta(days=1)))
        ws.cell(row=row_idx, column=4, value="present")
        ws.cell(row=row_idx, column=5, value="09:00:00")
        ws.cell(row=row_idx, column=6, value="18:00:00")

    buf = io.BytesIO()
    wb.save(buf)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=attendance_import_template.xlsx"}
    )


@router.post("/import/excel")
async def import_attendance_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin)
):
    import openpyxl
    import io
    from datetime import time

    try:
        contents = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
        ws = wb.active
        
        imported = 0
        errors = []
        
        def safe_date(val):
            if isinstance(val, (date, datetime)): return val
            if val is None or str(val).strip() == "": return None
            try:
                for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%Y/%m/%d"):
                    try: return datetime.strptime(str(val).strip(), fmt).date()
                    except: continue
                return None
            except: return None

        def safe_datetime(target_date, val):
            if isinstance(val, datetime): return val
            if isinstance(val, time): return datetime.combine(target_date, val)
            if val is None or str(val).strip() == "": return None
            val_str = str(val).strip()
            # Try parsing as time string
            for fmt in ("%H:%M:%S", "%H:%M", "%I:%M:%S %p", "%I:%M %p"):
                try:
                    t = datetime.strptime(val_str, fmt).time()
                    return datetime.combine(target_date, t)
                except: continue
            # Try parsing as full datetime string
            for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%d-%m-%Y %H:%M:%S", "%d/%m/%Y %H:%M:%S"):
                try:
                    return datetime.strptime(val_str, fmt)
                except: continue
            return None

        valid_statuses = ["present", "absent", "half_day", "weekly_off", "holiday", "leave", "on_duty", "late"]

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            # Skip completely empty rows
            if not row[0] or str(row[0]).strip() == "":
                continue
                
            emp_code = str(row[0]).strip()
            emp = db.query(HREmployee).filter(HREmployee.employee_id == emp_code).first()
            if not emp:
                errors.append(f"Row {row_idx}: Employee with ID '{emp_code}' not found")
                continue

            # Parse date
            record_date = safe_date(row[2])
            if not record_date:
                errors.append(f"Row {row_idx} ({emp.name}): Invalid date format")
                continue

            # Parse status
            status = str(row[3]).strip().lower() if row[3] else "present"
            if status not in valid_statuses:
                errors.append(f"Row {row_idx} ({emp.name}): Invalid status '{status}'")
                continue

            # Parse check-in / check-out
            check_in = safe_datetime(record_date, row[4]) if len(row) > 4 else None
            check_out = safe_datetime(record_date, row[5]) if len(row) > 5 else None

            try:
                # Find existing record or create new
                record = db.query(HRAttendanceRecord).filter(
                    HRAttendanceRecord.employee_id == emp.id,
                    HRAttendanceRecord.date == record_date
                ).first()

                if not record:
                    record = HRAttendanceRecord(employee_id=emp.id, date=record_date)
                    db.add(record)

                record.status = status
                record.check_in = check_in
                record.check_out = check_out
                record.correction_reason = "Imported via bulk Excel"
                record.corrected_by = current_user.id

                # Auto calculate hours worked
                hours_worked = 0
                if check_in and check_out:
                    hours_worked = (check_out - check_in).total_seconds() / 3600
                record.hours_worked = round(max(0.0, hours_worked), 2)

                db.flush()
                imported += 1
            except Exception as e:
                db.rollback()
                errors.append(f"Row {row_idx} ({emp.name}): {str(e)}")
                continue

        db.commit()
        return {"imported": imported, "errors": errors}
    except Exception as e:
        raise HTTPException(400, f"Critical failure parsing Excel file: {str(e)}")



