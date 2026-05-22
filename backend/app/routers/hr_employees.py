from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy.orm import Session
from pydantic import BaseModel
from typing import Optional, List
from datetime import date, datetime
from fastapi.responses import StreamingResponse
import openpyxl
import io
from app.database import get_db
from app.models import User
from app.auth import get_current_user
from app.hr_models import HREmployee, HRShift, HRLeaveBalance, HRLeaveType

router = APIRouter()

# ── Pydantic Schemas ──────────────────────────────────────────────────────────
class EmployeeCreate(BaseModel):
    employee_id: str
    name: str
    email: Optional[str] = None
    phone: Optional[str] = None
    designation: Optional[str] = None
    department_id: Optional[int] = None
    branch_id: Optional[int] = None
    manager_id: Optional[int] = None
    shift_id: Optional[int] = None
    date_of_joining: Optional[date] = None
    basic_salary: Optional[float] = 0
    salary_components: Optional[List] = []
    biometric_id: Optional[str] = None
    salary_template_id: Optional[int] = None
    user_id: Optional[int] = None
    salary_category: Optional[str] = "regular"
    enable_mobile_punch: Optional[bool] = False
    uan: Optional[str] = None
    esi_number: Optional[str] = None

class EmployeeUpdate(BaseModel):
    name: Optional[str] = None
    email: Optional[str] = None
    phone: Optional[str] = None
    designation: Optional[str] = None
    department_id: Optional[int] = None
    branch_id: Optional[int] = None
    manager_id: Optional[int] = None
    shift_id: Optional[int] = None
    date_of_joining: Optional[date] = None
    date_of_leaving: Optional[date] = None
    basic_salary: Optional[float] = None
    salary_components: Optional[List] = None
    biometric_id: Optional[str] = None
    salary_template_id: Optional[int] = None
    user_id: Optional[int] = None
    is_active: Optional[bool] = None
    salary_category: Optional[str] = None
    enable_mobile_punch: Optional[bool] = None
    uan: Optional[str] = None
    esi_number: Optional[str] = None

# ── Serializer ────────────────────────────────────────────────────────────────
def serialize(e: HREmployee):
    return {
        "id": e.id,
        "employee_id": e.employee_id,
        "user_id": e.user_id,
        "name": e.name,
        "email": e.email,
        "phone": e.phone,
        "designation": e.designation,
        "department_id": e.department_id,
        "department_name": e.department.name if e.department else None,
        "branch_id": e.branch_id,
        "branch_name": e.branch.name if e.branch else None,
        "manager_id": e.manager_id,
        "shift_id": e.shift_id,
        "shift_name": e.shift.name if e.shift else None,
        "date_of_joining": str(e.date_of_joining) if e.date_of_joining else None,
        "date_of_leaving": str(e.date_of_leaving) if e.date_of_leaving else None,
        "basic_salary": float(e.basic_salary or 0),
        "salary_components": e.salary_components or [],
        "biometric_id": e.biometric_id,
        "salary_template_id": e.salary_template_id,
        "salary_category": e.salary_category or "regular",
        "enable_mobile_punch": e.enable_mobile_punch or False,
        "uan": e.uan or "",
        "esi_number": e.esi_number or "",
        "is_active": e.is_active,
        "created_at": str(e.created_at),
    }

# ── Routes ────────────────────────────────────────────────────────────────────
@router.get("/")
def list_employees(
    branch_id: Optional[int] = None,
    department_id: Optional[int] = None,
    shift_id: Optional[int] = None,
    is_active: Optional[bool] = None,
    search: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    if not current_user.is_superadmin:
        from app.auth import get_current_employee_optional
        emp = get_current_employee_optional(current_user, db)
        if not emp:
            return []
        return [serialize(emp)]

    q = db.query(HREmployee)
    if branch_id:
        q = q.filter(HREmployee.branch_id == branch_id)
    if department_id:
        q = q.filter(HREmployee.department_id == department_id)
    if shift_id:
        q = q.filter(HREmployee.shift_id == shift_id)
    if is_active is not None:
        q = q.filter(HREmployee.is_active == is_active)
    if search:
        q = q.filter(
            HREmployee.name.ilike(f"%{search}%") |
            HREmployee.employee_id.ilike(f"%{search}%") |
            HREmployee.email.ilike(f"%{search}%")
        )
    return [serialize(e) for e in q.order_by(HREmployee.name).all()]


@router.post("/")
def create_employee(
    data: EmployeeCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    if db.query(HREmployee).filter(HREmployee.employee_id == data.employee_id).first():
        raise HTTPException(400, "Employee ID already exists")
    emp = HREmployee(**data.model_dump())
    db.add(emp); db.commit(); db.refresh(emp)
    return serialize(emp)


@router.get("/next-id")
def next_employee_id(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    last = db.query(HREmployee).order_by(HREmployee.id.desc()).first()
    if not last:
        return {"next_id": "EMP001"}
    try:
        num = int(last.employee_id.replace("EMP", "")) + 1
        return {"next_id": f"EMP{str(num).zfill(3)}"}
    except:
        return {"next_id": "EMP001"}


@router.get("/{emp_id}")
def get_employee(emp_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    emp = db.query(HREmployee).filter(HREmployee.id == emp_id).first()
    if not emp:
        raise HTTPException(404, "Employee not found")
    result = serialize(emp)
    # Include leave balances for this year
    year = datetime.now().year
    balances = db.query(HRLeaveBalance).filter(
        HRLeaveBalance.employee_id == emp_id,
        HRLeaveBalance.year == year
    ).all()
    result["leave_balances"] = [
        {
            "leave_type_id": b.leave_type_id,
            "leave_type_name": b.leave_type.name if b.leave_type else None,
            "leave_type_code": b.leave_type.code if b.leave_type else None,
            "allocated_days": b.allocated_days,
            "used_days": b.used_days,
            "remaining_days": b.allocated_days - b.used_days,
            "monthly_limit": b.monthly_limit or 0.0,
        }
        for b in balances
    ]
    return result


@router.put("/{emp_id}")
def update_employee(
    emp_id: int,
    data: EmployeeUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    emp = db.query(HREmployee).filter(HREmployee.id == emp_id).first()
    if not emp:
        raise HTTPException(404, "Employee not found")
    for k, v in data.model_dump(exclude_none=True).items():
        setattr(emp, k, v)
    db.commit(); db.refresh(emp)
    return serialize(emp)


@router.delete("/{emp_id}")
def deactivate_employee(emp_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    emp = db.query(HREmployee).filter(HREmployee.id == emp_id).first()
    if not emp:
        raise HTTPException(404, "Employee not found")
    emp.is_active = False
    db.commit()
    return {"message": "Employee deactivated"}

@router.get("/import/template")
def get_import_template(current_user: User = Depends(get_current_user)):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Employee Import Template"
    
    headers = [
        "employee_id", "name", "email", "phone", "designation", 
        "department_id", "branch_id", "shift_id", "date_of_joining", 
        "basic_salary", "biometric_id"
    ]
    
    # Style headers
    from openpyxl.styles import Font, PatternFill
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F46E5") # Primary accent color
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        
    # Add a sample row
    sample_row = [
        "EMP001", "John Doe", "john@example.com", "9876543210", "Manager",
        "", "", "", "2024-01-01", "50000", "101"
    ]
    for col_idx, val in enumerate(sample_row, 1):
        ws.cell(row=2, column=col_idx, value=val)

    # Add instructions in second sheet or comments
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['I'].width = 15
    
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    
    return StreamingResponse(
        buf, 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=employee_import_template.xlsx"}
    )

@router.post("/import/excel")
async def import_employees_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    try:
        contents = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
        ws = wb.active
        
        imported = 0
        skipped = 0
        errors = []
        
        def safe_int(val):
            if val is None or str(val).strip() == "": return None
            try: return int(float(val)) # handles "1.0" or 1
            except: return None

        def safe_float(val):
            if val is None or str(val).strip() == "": return 0.0
            try: return float(val)
            except: return 0.0

        def safe_date(val):
            if isinstance(val, (date, datetime)): return val
            if val is None or str(val).strip() == "": return None
            try:
                # Try common formats
                for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y"):
                    try: return datetime.strptime(str(val).strip(), fmt).date()
                    except: continue
                return None
            except: return None

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            # Row mapping: 0:id, 1:name, 2:email, 3:phone, 4:designation, 5:dept_id, 6:branch_id, 7:shift_id, 8:doj, 9:salary, 10:biometric_id
            if not row[0] or str(row[0]).strip() == "": 
                continue # Skip if no employee ID
            
            emp_id = str(row[0]).strip()
            name = str(row[1]).strip() if row[1] else None
            
            if not name:
                errors.append(f"Row {row_idx}: Name is missing")
                continue

            # Check if exists
            if db.query(HREmployee).filter(HREmployee.employee_id == emp_id).first():
                skipped += 1
                continue
                
            try:
                emp = HREmployee(
                    employee_id=emp_id,
                    name=name,
                    email=str(row[2]).strip() if row[2] else None,
                    phone=str(row[3]).strip() if row[3] else None,
                    designation=str(row[4]).strip() if row[4] else None,
                    department_id=safe_int(row[5]),
                    branch_id=safe_int(row[6]),
                    shift_id=safe_int(row[7]),
                    date_of_joining=safe_date(row[8]),
                    basic_salary=safe_float(row[9]),
                    biometric_id=str(row[10]).strip() if row[10] else None,
                    is_active=True
                )
                db.add(emp)
                db.flush() # Check for integrity errors (FK etc) immediately
                imported += 1
            except Exception as e:
                db.rollback()
                errors.append(f"Row {row_idx} ({emp_id}): {str(e)}")
                
        db.commit()
        return {"imported": imported, "skipped": skipped, "errors": errors}
    except Exception as e:
        raise HTTPException(400, f"Critical failure: {str(e)}")


